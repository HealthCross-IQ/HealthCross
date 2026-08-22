import io
from datetime import date

import pytest

from app.models import db_models as models


def _create_case(client, **overrides):
    payload = {"broker_name": "Broker", "company_name": "Amazonico", "industry": "trading"}
    payload.update(overrides)
    resp = client.post("/cases", json=payload)
    return resp.json()["id"]


def _insert_ledger_entries(client, case_id, months_and_amounts, policy_start=date(2025, 10, 1)):
    db = client.db_session_local()
    entry_id = 0
    for (year, month), amount in months_and_amounts.items():
        entry_id += 1
        db.add(
            models.ClaimsLedgerEntry(
                case_id=case_id,
                patient_id=f"P{entry_id}",
                claim_id=f"C{entry_id}",
                claim_status="Paid Claims",
                policy_start_date=policy_start,
                policy_end_date=date(policy_start.year + 1, policy_start.month, policy_start.day),
                date_of_treatment=date(year, month, 10),
                ip_op_maternity="OP",
                diagnosis_code="J209",
                diagnosis_description="Acute bronchitis",
                final_amount=amount,
            )
        )
    db.commit()
    db.close()


def test_case_update_sets_business_type_and_current_premium(client):
    case_id = _create_case(client)
    resp = client.patch(f"/cases/{case_id}", json={"business_type": "existing", "current_annual_premium": 3_000_000})
    assert resp.status_code == 200
    body = resp.json()
    assert body["business_type"] == "existing"
    assert body["current_annual_premium"] == 3_000_000


def test_claims_ledger_upload_replaces_not_accumulates(client):
    import io

    import pandas as pd

    case_id = _create_case(client)
    df = pd.DataFrame(
        [
            {
                "PATIENT_ID": "P1", "CLAIM_ID": "C1", "Claim Status": "Paid Claims",
                "POLICY_START_DATE": "2025-10-01", "POLICY_END_DATE": "2026-10-01",
                "DATE_OF_TREATMENT": "2025-11-05", "DIAGNOSIS_CODE": "J209",
                "DIAGNOSIS_DESCRIPTION": "Acute bronchitis", "Final Amount in AED": 500,
            }
        ]
    )
    buf = io.BytesIO()
    df.to_excel(buf, index=False)
    buf.seek(0)

    for _ in range(2):
        resp = client.post(
            f"/cases/{case_id}/claims-ledger",
            files={"file": ("ledger.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert resp.status_code == 200
        assert len(resp.json()) == 1


def test_claims_ledger_analysis_returns_top_patients_and_diagnoses(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C1", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 10, 5), diagnosis_code="C50", diagnosis_description="Malignant neoplasm of breast",
                ip_op_maternity="IP", final_amount=50000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C2", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 11, 5), diagnosis_code="Z511", diagnosis_description="Other medical care",
                ip_op_maternity="OP", final_amount=100,
            ),
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    body = resp.json()
    assert body["top_patients"][0]["patient_id"] == "P1"
    cancer = next(d for d in body["top_diagnoses"] if d["diagnosis_code"] == "C50")
    assert cancer["classification"] == "chronic"
    assert cancer["high_exposure"] is True


def test_claims_ledger_analysis_member_status_active_vs_deleted(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C1",
                policy_start_date=date(2025, 1, 1), policy_end_date=date(2025, 12, 31),
                member_start_date=date(2025, 1, 1), member_end_date=date(2025, 12, 31),
                date_of_treatment=date(2025, 10, 5), final_amount=5000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C2",
                policy_start_date=date(2025, 1, 1), policy_end_date=date(2025, 12, 31),
                member_start_date=date(2025, 1, 1), member_end_date=date(2025, 6, 30),
                date_of_treatment=date(2025, 11, 5), final_amount=100,
            ),
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    statuses = {p["patient_id"]: p["member_status"] for p in resp.json()["top_patients"]}
    assert statuses == {"P1": "Active", "P2": "Deleted"}


def test_claims_ledger_analysis_merges_monthly_erp_from_census(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            models.CensusRecord(
                case_id=case_id, employee_ref="E1",
                policy_start_date=date(2025, 10, 1), policy_end_date=date(2025, 11, 30),
                member_start_date=date(2025, 10, 1), member_end_date=date(2025, 11, 30),
            ),
            models.CensusRecord(
                case_id=case_id, employee_ref="E2",
                policy_start_date=date(2025, 10, 1), policy_end_date=date(2025, 11, 30),
                member_start_date=date(2025, 11, 1), member_end_date=date(2025, 11, 30),
            ),
        ]
    )
    db.commit()
    db.close()

    _insert_ledger_entries(client, case_id, {(2025, 10): 1000, (2025, 11): 2000, (2025, 12): 1500})

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    monthly = {(m["year"], m["month"]): m for m in resp.json()["monthly_final_amount"]}
    # October: only E1 covered (full month) -> ERP 1.0
    assert monthly[(2025, 10)]["erp"] == 1.0
    assert monthly[(2025, 10)]["cost_per_erp_member"] == 1000.0
    # November: E1 full month (1.0) + E2 full month (1.0) -> ERP 2.0
    assert monthly[(2025, 11)]["erp"] == 2.0
    assert monthly[(2025, 11)]["cost_per_erp_member"] == 1000.0


def test_claims_ledger_analysis_second_pmpm_based_annualization_method(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            # Member A present the whole term; Member B joins Nov 1 - ERP
            # grows from 1.0 (Oct) to 2.0 (Nov), so the plain monthly-total
            # average (Method 1) and the ERP-normalized rate (Method 2)
            # should diverge.
            models.CensusRecord(
                case_id=case_id, employee_ref="A",
                policy_start_date=date(2025, 10, 1), policy_end_date=date(2025, 12, 31),
                member_start_date=date(2025, 10, 1), member_end_date=date(2025, 12, 31),
            ),
            models.CensusRecord(
                case_id=case_id, employee_ref="B",
                policy_start_date=date(2025, 10, 1), policy_end_date=date(2025, 12, 31),
                member_start_date=date(2025, 11, 1), member_end_date=date(2025, 12, 31),
            ),
        ]
    )
    db.commit()
    db.close()

    _insert_ledger_entries(client, case_id, {(2025, 10): 1000, (2025, 11): 3000, (2025, 12): 500})

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    body = resp.json()

    # Method 1 (plain monthly-total average): (1000 + 3000) / 2 = 2000/month.
    assert body["avg_month"] == 2000.0
    assert body["annualized_incurred_claims"] == 24000.0

    # Method 2 (ERP-normalized rate): avg cost/member = (1000/1.0 + 3000/2.0) / 2 = 1250,
    # avg ERP = (1.0 + 2.0) / 2 = 1.5 -> annualized = 1250 * 12 * 1.5 = 22500.
    assert body["avg_cost_per_erp_member"] == 1250.0
    assert body["avg_erp"] == 1.5
    assert body["annualized_incurred_claims_pmpm"] == 22500.0
    assert body["annualized_incurred_claims_pmpm"] != body["annualized_incurred_claims"]


def test_claims_ledger_analysis_returns_top_providers(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C1", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 10, 5), provider_name="American Hospital", final_amount=5000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C2", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 11, 5), provider_name="Mediclinic City", final_amount=1000,
            ),
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    top_providers = resp.json()["top_providers"]
    assert top_providers[0]["provider_name"] == "American Hospital"
    assert top_providers[0]["final_amount"] == 5000.0


def test_claims_ledger_analysis_category_burning_cost_without_quote(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C1", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 10, 5), medical_category="A", final_amount=1000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C2", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 11, 5), medical_category="A", final_amount=3000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C3", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 12, 5), medical_category="A", final_amount=999999,
            ),  # trailing partial month - excluded
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    body = resp.json()
    assert body["quote_available_for_comparison"] is False
    category_row = next(r for r in body["category_burning_cost"] if r["category"] == "A")
    assert category_row["avg_month"] == 2000.0
    assert category_row["product"] is None
    assert category_row["projected_loss_ratio"] is None
    assert category_row["claim_count"] == 2  # C1, C2 - C3 excluded as the trailing partial month
    assert category_row["member_count"] == 1.0  # only P1 has claims within the full-months window
    assert category_row["avg_claims_per_member"] == category_row["projected_annual_claims"]
    assert category_row["pct_of_total_claims"] == 100.0  # only category present


def test_claims_ledger_analysis_category_burning_cost_matches_quoted_premium(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C1", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 10, 5), medical_category="A", final_amount=1000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C2", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 11, 5), medical_category="A", final_amount=3000,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C3", policy_start_date=date(2025, 10, 1),
                date_of_treatment=date(2025, 12, 5), medical_category="A", final_amount=999999,
            ),
        ]
    )
    db.add(
        models.BenefitPlan(
            case_id=case_id, role="quoted", category="A", plan_name="Gold Category A",
            network_type="MSH Platinum", gross_premium=50000.0, member_count=10,
        )
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 200
    body = resp.json()
    assert body["quote_available_for_comparison"] is True
    category_row = next(r for r in body["category_burning_cost"] if r["category"] == "A")
    assert category_row["product"] == "Gold Category A"
    assert category_row["network"] == "MSH Platinum"
    assert category_row["quoted_premium"] == 50000.0
    assert category_row["projected_loss_ratio"] == round(category_row["projected_annual_claims"] / 50000.0, 4)


def test_claims_ledger_analysis_404_without_upload(client):
    case_id = _create_case(client)
    resp = client.get(f"/cases/{case_id}/claims-ledger-analysis")
    assert resp.status_code == 404


def test_renewal_rating_end_to_end(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})

    monthly = {
        (2025, 10): 175385.01, (2025, 11): 155086.56, (2025, 12): 159281.85,
        (2026, 1): 311397.04, (2026, 2): 263984.21, (2026, 3): 433110.85,
        (2026, 4): 258319.43, (2026, 5): 305978.07, (2026, 6): 162401.95,
        (2026, 7): 57287.75,  # trailing partial month - must be excluded
    }
    _insert_ledger_entries(client, case_id, monthly)

    resp = client.get(f"/cases/{case_id}/renewal-rating")
    assert resp.status_code == 200
    body = resp.json()
    assert body["annualized_incurred_claims"] == 2966593.29
    assert body["current_annual_premium"] == 3_000_000.0
    assert len(body["months_used"]) == 9  # July excluded as trailing partial
    assert body["renewal_increase_pct"] > 0


def test_renewal_rating_requires_current_premium(client):
    case_id = _create_case(client)
    _insert_ledger_entries(client, case_id, {(2025, 10): 1000, (2025, 11): 2000, (2025, 12): 1500})
    resp = client.get(f"/cases/{case_id}/renewal-rating")
    assert resp.status_code == 400


def test_renewal_rating_credibility_style_dynamic_assumptions(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 100000, (2025, 11): 100000, (2025, 12): 100000, (2026, 1): 100000})

    default_resp = client.get(f"/cases/{case_id}/renewal-rating")
    lower_loading_resp = client.get(f"/cases/{case_id}/renewal-rating", params={"loading_pct": 0.10})
    assert lower_loading_resp.json()["renewal_increase_pct"] < default_resp.json()["renewal_increase_pct"]


def test_renewal_rating_includes_method_b_alongside_method_a(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 100000, (2025, 11): 100000, (2025, 12): 100000, (2026, 1): 100000})

    resp = client.get(f"/cases/{case_id}/renewal-rating")
    assert resp.status_code == 200
    body = resp.json()
    # Both methods share the SAME incurred-claims base (Paid+Outstanding+IBNR).
    assert body["annualized_incurred_claims"] == body["method_b"]["annualized_incurred_claims"]
    assert body["claims_with_ibnr"] == body["method_b"]["claims_with_ibnr"]
    assert body["assumptions_used"]["ibnr_pct"] == 0.10
    assert body["method_b"]["assumptions_used"]["ibnr_pct"] == 0.10
    # Method B (Burning Cost)'s credibility weighting means it requires LESS than Method A.
    assert body["method_b"]["required_premium"] < body["required_premium"]
    assert body["method_gap"] == round(body["method_b"]["required_premium"] - body["required_premium"], 2)


def test_renewal_rating_ibnr_pct_is_overridable(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 100000, (2025, 11): 100000, (2025, 12): 100000, (2026, 1): 100000})

    resp = client.get(f"/cases/{case_id}/renewal-rating", params={"ibnr_pct": 0.25})
    body = resp.json()
    assert body["assumptions_used"]["ibnr_pct"] == 0.25
    assert body["method_b"]["assumptions_used"]["ibnr_pct"] == 0.25


def test_renewal_rating_credibility_pct_is_overridable(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 100000, (2025, 11): 100000, (2025, 12): 100000, (2026, 1): 100000})

    resp = client.get(f"/cases/{case_id}/renewal-rating", params={"credibility_pct": 0.75})
    body = resp.json()
    assert body["method_b"]["assumptions_used"]["credibility_pct"] == 0.75
    assert body["assumptions_used"]["credibility_pct"] == 1.0  # Method A is never credibility-shaded


def test_renewal_benchmark_with_no_comparable_cases(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 80000, (2025, 11): 80000, (2025, 12): 80000})

    resp = client.get(f"/cases/{case_id}/renewal-benchmark")
    assert resp.status_code == 200
    body = resp.json()
    assert body["book"]["comparable_case_count"] == 0
    assert body["book"]["percentile"] is None
    assert body["book"]["low_credibility"] is True


def test_renewal_benchmark_compares_against_other_eligible_cases_only(client):
    this_case = _create_case(client, company_name="This Case")
    client.patch(f"/cases/{this_case}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, this_case, {(2025, 10): 90000, (2025, 11): 90000, (2025, 12): 90000})

    # A comparable case, eligible (ledger + premium).
    other_case = _create_case(client, company_name="Other Case")
    client.patch(f"/cases/{other_case}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, other_case, {(2025, 10): 40000, (2025, 11): 40000, (2025, 12): 40000})

    # An ineligible case (no current_annual_premium set) - must be skipped, not error the whole call.
    ineligible_case = _create_case(client, company_name="Ineligible Case")
    _insert_ledger_entries(client, ineligible_case, {(2025, 10): 50000, (2025, 11): 50000, (2025, 12): 50000})

    other_resp = client.get(f"/cases/{other_case}/renewal-rating")
    assert other_resp.status_code == 200
    other_loss_ratio = other_resp.json()["actual_loss_ratio"]

    resp = client.get(f"/cases/{this_case}/renewal-benchmark")
    assert resp.status_code == 200
    body = resp.json()
    assert body["book"]["comparable_case_count"] == 1  # only the eligible other case
    assert body["book"]["other_loss_ratios"] == [other_loss_ratio]
    assert body["book"]["percentile"] == 100.0  # this case's loss ratio (0.9) is higher
    assert body["book"]["low_credibility"] is True  # only 1 comparable case, below the threshold


def test_renewal_benchmark_404s_without_own_ledger(client):
    case_id = _create_case(client)
    resp = client.get(f"/cases/{case_id}/renewal-benchmark")
    assert resp.status_code == 404


def test_renewal_benchmark_400s_without_own_current_premium(client):
    case_id = _create_case(client)
    _insert_ledger_entries(client, case_id, {(2025, 10): 1000, (2025, 11): 2000, (2025, 12): 1500})
    resp = client.get(f"/cases/{case_id}/renewal-benchmark")
    assert resp.status_code == 400


def _upload_minimal_rate_card(client, tmp_path):
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "Product Name", "From Age", "To Age", "Male Price", "Female Price",
        "Married Female Price", "Region", "Network", "TPA", "Zone", "Created Date", "Updated Date",
    ])
    ws.append(["Gold", 0, 99, 3000, 3300, "0 (Applicable only for age band 18-50)", "Dubai", "Net A", "TPA X", "Worldwide", "2025-01-01", ""])
    path = tmp_path / "pricing.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        resp = client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200


def test_renewal_vs_new_business_generates_the_nb_premium_automatically(client, tmp_path):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 100000, (2025, 11): 100000, (2025, 12): 100000, (2026, 1): 100000})

    _upload_minimal_rate_card(client, tmp_path)
    db = client.db_session_local()
    db.add(models.CensusRecord(case_id=case_id, category="A", age=30, gender="M", marital_status="single", relation="employee", emirates="Dubai"))
    db.add(models.BenefitPlan(
        case_id=case_id, role="existing", plan_name="Category A", category="A",
        nb_product="Gold", nb_network="Net A", nb_tpa="TPA X", standard_summary={},
    ))
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/renewal-vs-new-business")
    assert resp.status_code == 200
    body = resp.json()
    assert body["new_business_module_premium"] == 4000.0  # single male, age 30, category A, Gold/Net A, net 3000 grossed up
    assert body["new_business_quote_id"] is not None
    assert body["gap"] == round(body["renewal_required_premium"] - 4000.0, 2)


def test_renewal_vs_new_business_returns_none_for_nb_side_without_a_rate_card(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 3_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 100000, (2025, 11): 100000, (2025, 12): 100000, (2026, 1): 100000})

    resp = client.get(f"/cases/{case_id}/renewal-vs-new-business")
    assert resp.status_code == 200
    body = resp.json()
    assert body["new_business_module_premium"] is None
    assert body["gap"] is None
    assert body["renewal_required_premium"] > 0


def test_renewal_vs_new_business_404s_without_own_ledger(client):
    case_id = _create_case(client)
    resp = client.get(f"/cases/{case_id}/renewal-vs-new-business")
    assert resp.status_code == 404


def test_case_large_claims_flags_a_one_off_claim_and_recurring_member(client):
    case_id = _create_case(client)
    db = client.db_session_local()
    db.add_all(
        [
            # One catastrophic claim dominating this case's own total.
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P1", claim_id="C1",
                date_of_treatment=date(2026, 1, 5), diagnosis_description="Cardiac surgery",
                claim_status="Paid Claims", final_amount=142000.0,
            ),
            # A member with 3 separate large-but-not-catastrophic claims.
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C2",
                date_of_treatment=date(2026, 1, 10), diagnosis_description="Dorsalgia",
                claim_status="Paid Claims", final_amount=55000.0,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C3",
                date_of_treatment=date(2026, 2, 10), diagnosis_description="Dorsalgia follow-up",
                claim_status="Paid Claims", final_amount=60000.0,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P2", claim_id="C4",
                date_of_treatment=date(2026, 3, 10), diagnosis_description="Dorsalgia follow-up 2",
                claim_status="Paid Claims", final_amount=52000.0,
            ),
            models.ClaimsLedgerEntry(
                case_id=case_id, patient_id="P3", claim_id="C5",
                date_of_treatment=date(2026, 1, 20), diagnosis_description="Allergic rhinitis",
                claim_status="Paid Claims", final_amount=500.0,
            ),
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/large-claims")
    assert resp.status_code == 200
    body = resp.json()

    assert body["top_claims"][0]["patient_id"] == "P1"
    assert body["top_claims"][0]["final_amount"] == 142000.0

    # 142,000 / (142,000+55,000+60,000+52,000+500) = ~45.6% of the total - well above the 15% one-off threshold.
    assert body["one_off_claim"] is not None
    assert body["one_off_claim"]["patient_id"] == "P1"
    assert body["one_off_claim"]["share_of_total_pct"] > 15

    recurring_ids = [m["patient_id"] for m in body["recurring_high_cost_members"]]
    assert recurring_ids == ["P2"]
    assert body["recurring_high_cost_members"][0]["large_claim_count"] == 3


def test_case_large_claims_404s_without_a_ledger(client):
    case_id = _create_case(client)
    resp = client.get(f"/cases/{case_id}/large-claims")
    assert resp.status_code == 404


def test_renewal_client_summary_full_case(client):
    case_id = _create_case(client, company_name="ServicePlan", broker_name="NASCO")
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 80000, (2025, 11): 80000, (2025, 12): 80000})

    db = client.db_session_local()
    db.add_all(
        [
            models.CensusRecord(case_id=case_id, age=30, gender="M", marital_status="single", relation="employee"),
            models.CensusRecord(case_id=case_id, age=28, gender="F", marital_status="married", relation="spouse"),
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/renewal-client-summary")
    assert resp.status_code == 200
    body = resp.json()

    assert body["case"]["company_name"] == "ServicePlan"
    assert body["case"]["broker_name"] == "NASCO"

    assert body["renewal"] is not None
    assert body["renewal"]["current_annual_premium"] == 1_000_000.0

    assert body["benchmark"] is not None
    assert body["benchmark"]["comparable_case_count"] == 0  # only case in the book

    assert body["premium_breakdown"] is not None
    proposed = body["premium_breakdown"]["proposed"]
    reconstructed = (
        proposed["risk_premium"] + proposed["tpa_fee"] + proposed["commission"]
        + proposed["hc_fee"] + proposed["qic_fee"]
    )
    assert reconstructed == pytest.approx(body["renewal"]["required_premium"], abs=0.5)

    assert body["census_summary"] is not None
    assert body["census_summary"]["total_members"] == 2

    assert body["top_diagnoses"] is not None
    assert body["top_diagnoses"][0]["diagnosis_description"] == "Acute bronchitis"


def test_renewal_client_summary_without_renewal_data_still_returns_case_and_census(client):
    case_id = _create_case(client, company_name="No Claims Yet")
    db = client.db_session_local()
    db.add(models.CensusRecord(case_id=case_id, age=40, gender="M", marital_status="single", relation="employee"))
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/renewal-client-summary")
    assert resp.status_code == 200
    body = resp.json()

    assert body["case"]["company_name"] == "No Claims Yet"
    assert body["renewal"] is None
    assert body["benchmark"] is None
    assert body["premium_breakdown"] is None
    assert body["top_diagnoses"] is None
    assert body["census_summary"]["total_members"] == 1


def test_renewal_client_summary_404s_for_missing_case(client):
    resp = client.get("/cases/999999/renewal-client-summary")
    assert resp.status_code == 404


def test_renewal_client_summary_uses_the_case_own_fee_split(client):
    case_id = _create_case(client)
    client.patch(
        f"/cases/{case_id}",
        json={
            "current_annual_premium": 1_000_000,
            "tpa_fee_pct": 0.15,
            "commission_pct": 0.10,
            "hc_fee_pct": 0.08,
            "qic_fee_pct": 0.05,
        },
    )
    _insert_ledger_entries(client, case_id, {(2025, 10): 80000, (2025, 11): 80000, (2025, 12): 80000})

    resp = client.get(f"/cases/{case_id}/renewal-client-summary")
    assert resp.status_code == 200
    breakdown = resp.json()["premium_breakdown"]
    loading_amount = breakdown["proposed"]["total"] - breakdown["proposed"]["risk_premium"]
    assert breakdown["proposed"]["tpa_fee"] == pytest.approx(loading_amount * 0.15 / 0.38, abs=0.5)


def test_renewal_client_summary_fee_split_changes_the_bottom_line_premium(client):
    """Regression test for the "loading breakdown isn't dynamic" bug: a
    case's own TPA Fee/Commission/HC Fee/QIC Fee % must feed into the
    actual required_premium/renewal_increase_pct, not just relabel a
    number still computed off the fixed default loading."""
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 80000, (2025, 11): 80000, (2025, 12): 80000})

    default_resp = client.get(f"/cases/{case_id}/renewal-client-summary")
    default_required_premium = default_resp.json()["renewal"]["required_premium"]

    client.patch(
        f"/cases/{case_id}",
        json={"tpa_fee_pct": 0.15, "commission_pct": 0.10, "hc_fee_pct": 0.08, "qic_fee_pct": 0.05},
    )
    custom_resp = client.get(f"/cases/{case_id}/renewal-client-summary")
    custom_body = custom_resp.json()
    custom_required_premium = custom_body["renewal"]["required_premium"]

    assert custom_required_premium != pytest.approx(default_required_premium)
    trended_claims = custom_body["renewal"]["trended_claims"]
    assert custom_required_premium == pytest.approx(trended_claims / (1 - 0.38), abs=0.5)


def _insert_census(client, case_id, members):
    db = client.db_session_local()
    db.add_all([models.CensusRecord(case_id=case_id, **m) for m in members])
    db.commit()
    db.close()


def test_member_rates_computes_new_rate_from_case_renewal_increase(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 80000, (2025, 11): 80000, (2025, 12): 80000})
    _insert_census(client, case_id, [{"employee_ref": "E1", "age": 30, "gender": "M", "relation": "employee"}])

    resp = client.get(f"/cases/{case_id}/renewal-client-summary")
    renewal_increase_pct = resp.json()["renewal"]["renewal_increase_pct"]

    get_resp = client.get(f"/cases/{case_id}/member-rates")
    assert get_resp.status_code == 200
    body = get_resp.json()
    assert body["case_renewal_increase_pct"] == pytest.approx(renewal_increase_pct)
    assert len(body["members"]) == 1
    census_record_id = body["members"][0]["census_record_id"]
    assert body["members"][0]["existing_annual_rate"] is None

    patch_resp = client.patch(
        f"/cases/{case_id}/member-rates",
        json=[{"census_record_id": census_record_id, "existing_annual_rate": 5000}],
    )
    assert patch_resp.status_code == 200
    member = patch_resp.json()["members"][0]
    assert member["existing_annual_rate"] == 5000
    expected_new_rate = round(5000 * (1 + renewal_increase_pct / 100), 2)
    assert member["computed_new_rate"] == pytest.approx(expected_new_rate)
    assert member["effective_new_rate"] == pytest.approx(expected_new_rate)
    assert member["rate_change_pct"] == pytest.approx(renewal_increase_pct, abs=0.01)


def test_member_rates_override_wins_over_computed_new_rate(client):
    case_id = _create_case(client)
    client.patch(f"/cases/{case_id}", json={"current_annual_premium": 1_000_000})
    _insert_ledger_entries(client, case_id, {(2025, 10): 80000, (2025, 11): 80000, (2025, 12): 80000})
    _insert_census(client, case_id, [{"employee_ref": "E1", "age": 30, "gender": "M", "relation": "employee"}])
    census_record_id = client.get(f"/cases/{case_id}/member-rates").json()["members"][0]["census_record_id"]

    patch_resp = client.patch(
        f"/cases/{case_id}/member-rates",
        json=[{"census_record_id": census_record_id, "existing_annual_rate": 5000, "new_annual_rate_override": 6000}],
    )
    member = patch_resp.json()["members"][0]
    assert member["computed_new_rate"] != 6000
    assert member["effective_new_rate"] == 6000
    assert member["rate_change_pct"] == pytest.approx(20.0)


def test_member_rates_without_renewal_data_leaves_new_rate_unset(client):
    case_id = _create_case(client, company_name="No Claims Yet")
    _insert_census(client, case_id, [{"employee_ref": "E1", "age": 40, "gender": "M", "relation": "employee"}])

    resp = client.get(f"/cases/{case_id}/member-rates")
    assert resp.status_code == 200
    body = resp.json()
    assert body["case_renewal_increase_pct"] is None
    assert body["members"][0]["computed_new_rate"] is None
    assert body["members"][0]["effective_new_rate"] is None


def test_member_rates_rejects_census_record_from_a_different_case(client):
    case_id = _create_case(client)
    other_case_id = _create_case(client, company_name="Other Co")
    _insert_census(client, other_case_id, [{"employee_ref": "E1", "age": 30, "gender": "M", "relation": "employee"}])
    other_census_id = client.get(f"/cases/{other_case_id}/member-rates").json()["members"][0]["census_record_id"]

    resp = client.patch(
        f"/cases/{case_id}/member-rates",
        json=[{"census_record_id": other_census_id, "existing_annual_rate": 1000}],
    )
    assert resp.status_code == 404


def test_member_rates_404s_for_missing_case(client):
    resp = client.get("/cases/999999/member-rates")
    assert resp.status_code == 404


def _rate_card_xlsx_bytes():
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None, None, None, None, None, None, None, None, "Brokerage in %", 0.125])
    ws.append([None, None, None, None, None, None, None, None, None, "TPA Fee in AED", 0.065])
    ws.append([None, None, None, None, None, None, None, None, None, "Health CROSS", 0.065])
    ws.append(["Category", "Age Band", "Gross Premium in AED"])
    ws.append(["Category A Male", "0-17", 9117])
    ws.append(["Category A Male", "18-40", 11664])
    ws.append(["Category A Female", "0-17", 9116])
    ws.append(["Category A Female", "18-40", 13611])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_import_member_rate_card_fills_existing_rates_by_category_gender_age(client):
    case_id = _create_case(client)
    _insert_census(client, case_id, [
        {"employee_ref": "E1", "category": "A", "age": 30, "gender": "M", "relation": "employee"},
        {"employee_ref": "E2", "category": "A", "age": 10, "gender": "F", "relation": "child"},
        {"employee_ref": "E3", "category": "B", "age": 30, "gender": "M", "relation": "employee"},  # no Category B rows
    ])

    resp = client.post(
        f"/cases/{case_id}/member-rates/import-rate-card",
        files={"file": ("rates.xlsx", _rate_card_xlsx_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 200
    body = resp.json()

    assert body["matched_count"] == 2
    assert len(body["unmatched"]) == 1
    assert body["unmatched"][0]["employee_ref"] == "E3"
    assert body["detected_fees"] == {
        "commission_pct": pytest.approx(0.125),
        "tpa_fee_pct": pytest.approx(0.065),
        "hc_fee_pct": pytest.approx(0.065),
    }

    members_by_ref = {m["employee_ref"]: m for m in body["members"]}
    assert members_by_ref["E1"]["existing_annual_rate"] == 11664.0
    assert members_by_ref["E2"]["existing_annual_rate"] == 9116.0
    assert members_by_ref["E3"]["existing_annual_rate"] is None


def test_import_member_rate_card_404s_for_missing_case(client):
    resp = client.post(
        "/cases/999999/member-rates/import-rate-card",
        files={"file": ("rates.xlsx", _rate_card_xlsx_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 404


def test_import_member_rate_card_400s_for_a_file_with_no_rate_table(client):
    case_id = _create_case(client)
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Nothing", "Relevant", "Here"])
    buf = io.BytesIO()
    wb.save(buf)

    resp = client.post(
        f"/cases/{case_id}/member-rates/import-rate-card",
        files={"file": ("rates.xlsx", buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert resp.status_code == 400
