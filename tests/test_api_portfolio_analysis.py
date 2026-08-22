"""API tests for Portfolio Analysis (app/api/routes_portfolio_analysis.py) -
uploading the book's own membership/claims/group-product-mapping data and
running the analysis against a rate card. Uses small synthetic
spreadsheets rather than HealthCross's real (client PII) book export.
"""
import openpyxl
import pytest

MEMBERS_HEADER = [
    "CONTRACT", "MASTERCONTRACT", "POLICYNUMBER", "MSH_POLICYNUMBER", "BENEFICIARYID",
    "DOB", "GENDER", "MARITALSTATUS", "NATIONALITY", "DEPENDENCY",
    "PERSONRESIDENCEEMIRATE", "CATEGORY", "NETWORKTYPE",
    "Eff Date", "Exp Date", "EndoDate (Member Start Date)", "EndoDate (Member End Date)",
    "GrossPremium", "ActualGrossPremium", "NETPREMIUM", "ACTUALNETPREMIUM", "TPA FEE",
]

CLAIMS_HEADER = [
    "PATIENT_ID", "CLAIM_ID", "Claim Status", "GROUP_NAME", "CLIENT_NAME", "MSH_POLICY_NUMBER",
    "POLICY_START_DATE", "POLICY_END_DATE", "Member Start Date", "Member End Date",
    "DATE_OF_TREATMENT", "RELATION", "IP_OP_MATERNITY", "MEDICAL_CATEGORY", "PROVIDER_NAME",
    "DIAGNOSIS_CODE", "DIAGNOSIS_DESCRIPTION", "Claimed Amount AED", "Final Amount in AED",
]

PRODUCT_PRICING_HEADER = [
    "Product Name", "From Age", "To Age", "Male Price", "Female Price",
    "Married Female Price", "Region", "Network", "TPA", "Zone", "Created Date", "Updated Date",
]


def _write_xlsx(tmp_path, name, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


@pytest.fixture()
def members_xlsx(tmp_path):
    return _write_xlsx(
        tmp_path,
        "members.xlsx",
        MEMBERS_HEADER,
        [
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                12000, 12000, None, None, 500,
            ],
        ],
    )


@pytest.fixture()
def claims_xlsx(tmp_path):
    return _write_xlsx(
        tmp_path,
        "claims.xlsx",
        CLAIMS_HEADER,
        [
            [
                "ACM0001", "CLM1", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                "2025-06-01", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
                "J309", "Allergic rhinitis", 100.0, 90.0,
            ],
        ],
    )


@pytest.fixture()
def mapping_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Master Group", "Product"])
    ws.append(["Acme Sub LLC", "Acme Holdings", "Gold"])
    path = tmp_path / "mapping.xlsx"
    wb.save(path)
    return path


@pytest.fixture()
def members_two_policy_years_xlsx(tmp_path):
    # Same client, two members on different policy years - simulates a
    # group that's already renewed, with last year's and this year's
    # members both showing up in the same current-book export.
    return _write_xlsx(
        tmp_path,
        "members_two_years.xlsx",
        MEMBERS_HEADER,
        [
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                12000, 12000, None, None, 500,
            ],
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-B", "ACM0002",
                "1985-06-01", "F", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2026-01-01", "2027-01-01", "2026-01-01", "2027-01-01",
                13000, 13000, None, None, 500,
            ],
        ],
    )


@pytest.fixture()
def members_two_subgroups_same_master_xlsx(tmp_path):
    # Two different subgroups (contracts) under the same master policy -
    # lets a test verify master_client rolls both up into one row while
    # client (subgroup) still shows them separately.
    return _write_xlsx(
        tmp_path,
        "members_two_subgroups.xlsx",
        MEMBERS_HEADER,
        [
            [
                "Acme Sub A", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                12000, 12000, None, None, 500,
            ],
            [
                "Acme Sub B", "Acme Holdings", "P200", "QC1-ACM-B", "ACM0002",
                "1988-01-01", "F", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/B", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                13000, 13000, None, None, 500,
            ],
        ],
    )


@pytest.fixture()
def rate_card_xlsx(tmp_path):
    return _write_xlsx(
        tmp_path,
        "pricing.xlsx",
        PRODUCT_PRICING_HEADER,
        [
            ["Gold", 18, 40, 4000, 4400, "0 (Applicable only for age band 18-50)", "Dubai", "MSH Platinum", "MSH MENA", "Worldwide", "2025-01-01", ""],
        ],
    )


@pytest.fixture()
def members_two_networks_xlsx(tmp_path):
    # Same client and Product, two different networks - lets a test stack
    # a product filter with a network filter to isolate just one slice.
    return _write_xlsx(
        tmp_path,
        "members_two_networks.xlsx",
        MEMBERS_HEADER,
        [
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                12000, 12000, None, None, 500,
            ],
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-B", "ACM0002",
                "1988-01-01", "F", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "COMPREHENSIVE",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                9000, 9000, None, None, 500,
            ],
        ],
    )


@pytest.fixture()
def members_unreliable_master_field_xlsx(tmp_path):
    # Real system exports sometimes leave MASTERCONTRACT blank/wrong (here,
    # it just duplicates each row's own subgroup) - the uploaded
    # Group->Product mapping file is the only reliable source for the true
    # subgroup->master relationship in that case.
    return _write_xlsx(
        tmp_path,
        "members_unreliable_master.xlsx",
        MEMBERS_HEADER,
        [
            [
                "Acme Sub A", "Acme Sub A", "P100", "QC1-ACM-A", "ACM0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                12000, 12000, None, None, 500,
            ],
            [
                "Acme Sub B", "Acme Sub B", "P200", "QC1-ACM-B", "ACM0002",
                "1988-01-01", "F", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/B", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                13000, 13000, None, None, 500,
            ],
        ],
    )


@pytest.fixture()
def master_mapping_for_unreliable_xlsx(tmp_path):
    # The dedicated Subgroup -> Master Group mapping sheet - a plain
    # two-column file, distinct from the Group->Product mapping.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Group Name"])
    ws.append(["Acme Sub A", "Acme Holdings"])
    ws.append(["Acme Sub B", "Acme Holdings"])
    path = tmp_path / "master_mapping.xlsx"
    wb.save(path)
    return path


@pytest.fixture()
def renewed_member_two_years_xlsx(tmp_path):
    # The SAME beneficiary (ACQ0001) renewed - appears as two rows, one
    # per policy year, sharing one beneficiary ID.
    return _write_xlsx(
        tmp_path,
        "renewed_member.xlsx",
        MEMBERS_HEADER,
        [
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACQ0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                10000, 10000, None, None, 500,
            ],
            [
                "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACQ0001",
                "1990-01-01", "M", "Single", "India", "Principal",
                "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
                "2026-01-01", "2027-01-01", "2026-01-01", "2027-01-01",
                11000, 11000, None, None, 500,
            ],
        ],
    )


@pytest.fixture()
def renewed_member_claims_xlsx(tmp_path):
    # One claim dated in each policy year, both against the same patient.
    return _write_xlsx(
        tmp_path,
        "renewed_member_claims.xlsx",
        CLAIMS_HEADER,
        [
            [
                "ACQ0001", "CLM1", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
                "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
                "2025-06-01", "Main Insured", "OP", "GP", "Clinic X",
                "Z00", "Checkup", 1000.0, 900.0,
            ],
            [
                "ACQ0001", "CLM2", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
                "2026-01-01", "2027-01-01", "2026-01-01", "2027-01-01",
                "2026-06-01", "Main Insured", "IP", "SURGERY", "Hospital Y",
                "K358", "Appendicitis", 5000.0, 4800.0,
            ],
        ],
    )


def test_upload_members_ingests_rows(client, members_xlsx):
    with open(members_xlsx, "rb") as f:
        resp = client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200
    assert resp.json()["rows_ingested"] == 1


def test_upload_members_replaces_existing_rows(client, members_xlsx):
    for _ in range(2):
        with open(members_xlsx, "rb") as f:
            resp = client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    assert resp.json()["rows_ingested"] == 1


def test_upload_claims_ingests_rows(client, claims_xlsx):
    with open(claims_xlsx, "rb") as f:
        resp = client.post("/portfolio-analysis/claims/upload", files={"file": ("claims.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200
    assert resp.json()["rows_ingested"] == 1


@pytest.fixture()
def large_claims_xlsx(tmp_path):
    return _write_xlsx(
        tmp_path,
        "large_claims.xlsx",
        CLAIMS_HEADER,
        [
            # ACM0001: one single catastrophic claim - not "recurring".
            ["ACM0001", "CLM1", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
             "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
             "2025-06-01", "Main Insured", "IP", "HOSPITALISATION", "Big Hospital",
             "C500", "Cancer treatment", 300000.0, 300000.0],
            # ACM0002: three separate large claims - recurring high-cost.
            ["ACM0002", "CLM2", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
             "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
             "2025-03-01", "Main Insured", "IP", "HOSPITALISATION", "Some Hospital",
             "K358", "Appendicitis", 60000.0, 60000.0],
            ["ACM0002", "CLM3", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
             "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
             "2025-05-01", "Main Insured", "IP", "HOSPITALISATION", "Some Hospital",
             "K358", "Appendicitis follow-up", 55000.0, 55000.0],
            ["ACM0002", "CLM4", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
             "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
             "2025-07-01", "Main Insured", "IP", "HOSPITALISATION", "Some Hospital",
             "K358", "Appendicitis follow-up 2", 70000.0, 70000.0],
            # A small, ordinary claim - shouldn't show up in any large-claims view.
            ["ACM0003", "CLM5", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
             "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
             "2025-06-01", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
             "J309", "Allergic rhinitis", 100.0, 90.0],
        ],
    )


def test_large_claims_404s_without_any_claims_uploaded(client):
    resp = client.get("/portfolio-analysis/large-claims")
    assert resp.status_code == 400


def test_large_claims_returns_top_claims_top_members_thresholds_and_recurring(client, large_claims_xlsx):
    with open(large_claims_xlsx, "rb") as f:
        client.post("/portfolio-analysis/claims/upload", files={"file": ("large_claims.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/large-claims")
    assert resp.status_code == 200
    body = resp.json()

    assert body["top_claims"][0]["patient_id"] == "ACM0001"
    assert body["top_claims"][0]["final_amount"] == 300000.0

    top_member_ids = [m["patient_id"] for m in body["top_members"]]
    # ACM0001's single 300,000 claim still beats ACM0002's summed 185,000
    # across three claims - top_members ranks by TOTAL, not claim count.
    assert top_member_ids[0] == "ACM0001"
    assert body["top_members"][0]["total_claims"] == 300000.0
    assert top_member_ids[1] == "ACM0002"
    assert body["top_members"][1]["total_claims"] == 185000.0
    assert body["top_members"][1]["claim_count"] == 3

    buckets = {b["threshold"]: b for b in body["threshold_buckets"]}
    assert buckets[50000.0]["claim_count"] == 4  # every large-claim line above (300k, 60k, 55k, 70k)
    assert buckets[250000.0]["claim_count"] == 1

    recurring_ids = [m["patient_id"] for m in body["recurring_high_cost_members"]]
    assert recurring_ids == ["ACM0002"]
    assert body["recurring_high_cost_members"][0]["large_claim_count"] == 3


def test_upload_group_product_mapping_ingests_rows(client, mapping_xlsx):
    with open(mapping_xlsx, "rb") as f:
        resp = client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200
    assert resp.json()["rows_ingested"] == 2  # subgroup + master group both get a row


def test_upload_subgroup_mapping_ingests_rows(client, master_mapping_for_unreliable_xlsx):
    with open(master_mapping_for_unreliable_xlsx, "rb") as f:
        resp = client.post("/portfolio-analysis/subgroup-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200
    assert resp.json()["rows_ingested"] == 2


def test_upload_subgroup_mapping_replaces_existing_rows(client, master_mapping_for_unreliable_xlsx):
    for _ in range(2):
        with open(master_mapping_for_unreliable_xlsx, "rb") as f:
            resp = client.post("/portfolio-analysis/subgroup-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    assert resp.json()["rows_ingested"] == 2


def test_summary_requires_members_uploaded_first(client):
    resp = client.get("/portfolio-analysis/summary")
    assert resp.status_code == 400


def test_summary_requires_rate_card_uploaded(client, members_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    resp = client.get("/portfolio-analysis/summary")
    assert resp.status_code == 400


def test_full_pipeline_summary_by_product(client, members_xlsx, claims_xlsx, mapping_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(claims_xlsx, "rb") as f:
        client.post("/portfolio-analysis/claims/upload", files={"file": ("claims.xlsx", f, "application/octet-stream")})
    with open(mapping_xlsx, "rb") as f:
        client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "product"})
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_members"] == 1
    assert body["out_of_scope_member_count"] == 0
    assert body["unmapped_product_member_count"] == 0

    gold = next(r for r in body["rows"] if r["product"] == "Gold")
    assert gold["member_count"] == 1
    assert gold["standard_premium"] == 4000.0  # male, 18-40, Gold/MSH Platinum
    assert gold["actual_premium"] == 12000.0
    assert gold["actual_claims"] == 90.0


def test_summary_rejects_an_invalid_group_by(client, members_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "not_a_field"})
    assert resp.status_code == 400


def test_member_detail_endpoint_returns_per_member_rows(client, members_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/members")
    assert resp.status_code == 200
    body = resp.json()
    assert len(body) == 1
    assert body[0]["beneficiary_id"] == "ACM0001"


def test_data_as_of_defaults_to_none(client):
    resp = client.get("/portfolio-analysis/data-as-of")
    assert resp.status_code == 200
    assert resp.json()["data_as_of_date"] is None


def test_set_data_as_of_directly(client):
    resp = client.post("/portfolio-analysis/data-as-of", json={"data_as_of_date": "2026-07-15"})
    assert resp.status_code == 200
    assert resp.json()["data_as_of_date"] == "2026-07-15"

    resp = client.get("/portfolio-analysis/data-as-of")
    assert resp.json()["data_as_of_date"] == "2026-07-15"


def test_upload_members_with_data_as_of_form_field_sets_the_stored_date(client, members_xlsx):
    with open(members_xlsx, "rb") as f:
        resp = client.post(
            "/portfolio-analysis/members/upload",
            files={"file": ("members.xlsx", f, "application/octet-stream")},
            data={"data_as_of": "2025-07-01"},
        )
    assert resp.status_code == 200
    resp = client.get("/portfolio-analysis/data-as-of")
    assert resp.json()["data_as_of_date"] == "2025-07-01"


def test_summary_defaults_earned_premium_to_the_stored_as_of_date(client, members_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    # No stored/explicit as_of yet - the member's policy (2025-01-01 to
    # 2026-01-01) has already ended by "today", so fully earned.
    resp = client.get("/portfolio-analysis/summary", params={"group_by": "network"})
    full = next(r for r in resp.json()["rows"] if r["network"] == "MSH Platinum")
    assert full["actual_premium"] == 12000.0

    # Set a stored as-of date partway through that policy period - earned
    # premium should drop below the full annual amount without needing an
    # explicit as_of on the request itself.
    client.post("/portfolio-analysis/data-as-of", json={"data_as_of_date": "2025-04-02"})
    resp = client.get("/portfolio-analysis/summary", params={"group_by": "network"})
    partial = next(r for r in resp.json()["rows"] if r["network"] == "MSH Platinum")
    assert partial["actual_premium"] < 12000.0

    # An explicit as_of query param overrides the stored date.
    resp = client.get("/portfolio-analysis/summary", params={"group_by": "network", "as_of": "2026-06-01"})
    overridden = next(r for r in resp.json()["rows"] if r["network"] == "MSH Platinum")
    assert overridden["actual_premium"] == 12000.0


def test_summary_group_by_policy_year_separates_renewal_cohorts(client, members_two_policy_years_xlsx, rate_card_xlsx):
    with open(members_two_policy_years_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "policy_year"})
    assert resp.status_code == 200
    rows = {r["policy_year"]: r for r in resp.json()["rows"]}
    assert rows["2025"]["member_count"] == 1
    assert rows["2026"]["member_count"] == 1


def test_summary_policy_year_filter_narrows_to_one_cohort(client, members_two_policy_years_xlsx, rate_card_xlsx):
    with open(members_two_policy_years_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "client", "policy_year": "2026"})
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_members"] == 1
    acme = next(r for r in body["rows"] if r["client"] == "Acme Sub LLC")
    assert acme["member_count"] == 1


def test_summary_by_policy_year_does_not_double_count_a_renewed_members_claims(
    client, renewed_member_two_years_xlsx, renewed_member_claims_xlsx, rate_card_xlsx,
):
    with open(renewed_member_two_years_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(renewed_member_claims_xlsx, "rb") as f:
        client.post("/portfolio-analysis/claims/upload", files={"file": ("claims.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "policy_year"})
    assert resp.status_code == 200
    rows = {r["policy_year"]: r for r in resp.json()["rows"]}
    # The same beneficiary ID (ACQ0001) has one row per policy year - each
    # year's claims must only count once, against its own year, not both.
    assert rows["2025"]["actual_claims"] == 900.0
    assert rows["2026"]["actual_claims"] == 4800.0


def test_summary_policy_year_filter_with_no_matches_returns_400(client, members_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "product", "policy_year": "1999"})
    assert resp.status_code == 400


def test_burning_cost_by_age_gender_endpoint(client, members_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/burning-cost-by-age-gender")
    assert resp.status_code == 200
    rows = resp.json()
    assert any(r["age_band"] == "18-40" and r["gender"] == "M" for r in rows)


def test_insights_endpoint_returns_every_breakdown_from_one_call(client, members_xlsx, rate_card_xlsx):
    # The Portfolio Insights dashboard used to make 7 separate requests
    # (one per group_by dimension, plus age/gender), each re-running the
    # full member/claims analysis from scratch - this single endpoint
    # returns every one of those breakdowns from one shared analysis run.
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/insights")
    assert resp.status_code == 200
    data = resp.json()
    assert set(data.keys()) == {
        "by_product", "by_network", "by_nationality_zone", "by_relation", "by_gender", "by_policy_year",
        "by_category", "by_subgroup", "by_age_gender",
    }
    assert data["by_product"]["group_by"] == "product"
    assert data["by_product"]["total_members"] == data["by_network"]["total_members"]
    assert any(r["age_band"] == "18-40" and r["gender"] == "M" for r in data["by_age_gender"])

    assert data["by_category"]["group_by"] == "category"
    assert data["by_category"]["rows"][0]["category"] == "QIC/HC/BR/ACM/DXB/A"
    assert data["by_category"]["rows"][0]["network"] == "MSH Platinum"

    assert data["by_subgroup"]["group_by"] == "client"
    assert data["by_subgroup"]["rows"][0]["client"] == "Acme Sub LLC"
    assert data["by_subgroup"]["rows"][0]["policy_start_date"] == "2025-01-01"


def test_demographic_summary_endpoint_matches_insights_total_members(client, members_xlsx, mapping_xlsx, rate_card_xlsx):
    # Regression coverage for a real question raised while reviewing a
    # mockup: the Demographic view's headline total must match whatever
    # every other Portfolio Analysis view (Insights, Summary/"Loss ratio by
    # Product") already reports for the same book, not a smaller,
    # inconsistent-looking number from silently excluding out-of-scope
    # members.
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(mapping_xlsx, "rb") as f:
        client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/insights")
    insights_total = resp.json()["by_product"]["total_members"]

    resp = client.get("/portfolio-analysis/demographic-summary")
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_members"] == insights_total
    assert body["gender_counts"]["M"] == 1
    assert body["product_counts"] == {"Gold": 1}
    assert body["network_counts"] == {"MSH Platinum": 1}
    assert body["nationality_zone_top5"]["zone_1_asia"] == [{"nationality": "India", "count": 1}]


def test_insights_endpoint_respects_the_master_client_filter(client, members_two_policy_years_xlsx, rate_card_xlsx):
    with open(members_two_policy_years_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    # No Subgroup->Master mapping uploaded here, so master_client resolves
    # to the raw master_contract field ("Acme Holdings"), not the subgroup
    # name itself - see resolve_master_client's fallback order.
    resp = client.get("/portfolio-analysis/insights", params={"master_client": "Acme Holdings"})
    assert resp.status_code == 200
    assert resp.json()["by_product"]["total_members"] == 2

    resp = client.get("/portfolio-analysis/insights", params={"master_client": "Nonexistent Co"})
    assert resp.status_code == 200
    assert resp.json()["by_product"]["total_members"] == 0


def test_burning_cost_by_product_network_endpoint(client, members_xlsx, mapping_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(mapping_xlsx, "rb") as f:
        client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/burning-cost-by-product-network")
    assert resp.status_code == 200
    rows = resp.json()
    assert any(r["product"] == "Gold" and r["network"] == "MSH Platinum" for r in rows)


def test_burning_cost_by_product_network_endpoint_returns_empty_list_without_a_rate_card(client):
    # Optional supporting context for New Business Rating, not something
    # that page requires - a 400 here would be a red herring on a page
    # whose own required uploads are unrelated to Portfolio Analysis.
    resp = client.get("/portfolio-analysis/burning-cost-by-product-network")
    assert resp.status_code == 200
    assert resp.json() == []


def test_burning_cost_by_product_network_age_gender_endpoint(client, members_xlsx, mapping_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(mapping_xlsx, "rb") as f:
        client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/burning-cost-by-product-network-age-gender")
    assert resp.status_code == 200
    rows = resp.json()
    assert any(r["product"] == "Gold" and r["network"] == "MSH Platinum" and r["age_band"] == "18-40" for r in rows)


def test_burning_cost_by_product_network_age_gender_endpoint_returns_empty_list_without_a_rate_card(client):
    resp = client.get("/portfolio-analysis/burning-cost-by-product-network-age-gender")
    assert resp.status_code == 200
    assert resp.json() == []


def test_list_clients_endpoint_returns_distinct_contract_names(client, members_two_policy_years_xlsx):
    with open(members_two_policy_years_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/clients")
    assert resp.status_code == 200
    assert resp.json() == ["Acme Sub LLC"]


def test_summary_client_filter_scopes_to_one_client(client, members_two_policy_years_xlsx, rate_card_xlsx):
    with open(members_two_policy_years_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "policy_year", "client": "Acme Sub LLC"})
    assert resp.status_code == 200
    assert resp.json()["total_members"] == 2


def test_summary_stacks_product_and_network_filters_together(client, members_two_networks_xlsx, mapping_xlsx, rate_card_xlsx):
    with open(members_two_networks_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(mapping_xlsx, "rb") as f:
        client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    # Both members are Gold, but on different networks - stacking both
    # filters narrows to just the Platinum-network member.
    resp = client.get(
        "/portfolio-analysis/summary",
        params={"group_by": "gender", "product": "Gold", "network": "MSH Platinum"},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_members"] == 1
    assert body["rows"][0]["gender"] == "M"


def test_summary_client_filter_with_unknown_client_returns_400(client, members_xlsx, rate_card_xlsx):
    with open(members_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "product", "client": "Nonexistent Co"})
    assert resp.status_code == 400


def test_list_master_clients_combines_subgroups_under_one_master(client, members_two_subgroups_same_master_xlsx):
    with open(members_two_subgroups_same_master_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/master-clients")
    assert resp.status_code == 200
    assert resp.json() == ["Acme Holdings"]

    # But the subgroup picker still shows both separately.
    resp = client.get("/portfolio-analysis/clients")
    assert sorted(resp.json()) == ["Acme Sub A", "Acme Sub B"]


def test_summary_group_by_master_client_combines_subgroups(client, members_two_subgroups_same_master_xlsx, rate_card_xlsx):
    with open(members_two_subgroups_same_master_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "master_client"})
    assert resp.status_code == 200
    body = resp.json()
    assert len(body["rows"]) == 1
    assert body["rows"][0]["master_client"] == "Acme Holdings"
    assert body["rows"][0]["member_count"] == 2


def test_summary_master_client_filter_then_group_by_client_shows_subgroups(client, members_two_subgroups_same_master_xlsx, rate_card_xlsx):
    with open(members_two_subgroups_same_master_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get(
        "/portfolio-analysis/summary",
        params={"group_by": "client", "master_client": "Acme Holdings"},
    )
    assert resp.status_code == 200
    subgroups = {r["client"] for r in resp.json()["rows"]}
    assert subgroups == {"Acme Sub A", "Acme Sub B"}


def test_master_clients_use_the_uploaded_mapping_not_the_unreliable_raw_field(
    client, members_unreliable_master_field_xlsx, master_mapping_for_unreliable_xlsx, rate_card_xlsx,
):
    with open(members_unreliable_master_field_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    # Before the mapping is uploaded, MASTERCONTRACT is unreliable (just
    # duplicates each subgroup), so the two subgroups look like two
    # separate masters.
    resp = client.get("/portfolio-analysis/master-clients")
    assert sorted(resp.json()) == ["Acme Sub A", "Acme Sub B"]

    with open(master_mapping_for_unreliable_xlsx, "rb") as f:
        client.post("/portfolio-analysis/subgroup-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})

    # Once uploaded, both subgroups correctly resolve to their real master.
    resp = client.get("/portfolio-analysis/master-clients")
    assert resp.json() == ["Acme Holdings"]

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "master_client"})
    assert resp.status_code == 200
    rows = resp.json()["rows"]
    assert len(rows) == 1
    assert rows[0]["master_client"] == "Acme Holdings"
    assert rows[0]["member_count"] == 2


def test_master_clients_combine_many_subgroups_under_one_real_master_name(client, rate_card_xlsx):
    # The second real-world example the user shared: MPH's 6 subgroups
    # (including the master's own name as a subgroup row) must all
    # resolve to exactly one master-client name in the dropdown, and
    # their combined loss ratio/burning cost must roll up together.
    members_wb = openpyxl.Workbook()
    ws = members_wb.active
    ws.append(MEMBERS_HEADER)
    subgroups = [
        "MPH CONSULTING SERVICES DMCC", "MPH ENGINEERING SERVICES QFC", "MPH GLOBAL SERVICES DMCC",
        "MPH ON DEMAND LABORS SUPPLY", "MPH RECRUITMENT SERVICES", "MPH TECHNICAL SERVICES W.L.L",
    ]
    for i, subgroup in enumerate(subgroups):
        ws.append([
            subgroup, subgroup, f"P{i}", f"QC1-{i}", f"MPH{i:04d}",
            "1990-01-01", "M", "Single", "India", "Principal",
            "Dubai", "CAT1", "PLATINUM",
            "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
            10000, 10000, None, None, 500,
        ])
    members_path_dir = rate_card_xlsx.parent
    members_path = members_path_dir / "mph_members.xlsx"
    members_wb.save(members_path)

    mapping_wb = openpyxl.Workbook()
    ws2 = mapping_wb.active
    ws2.append(["Subgroup", "Group Name"])
    for subgroup in subgroups:
        ws2.append([subgroup, "MPH CONSULTING SERVICES DMCC"])
    mapping_path = members_path_dir / "mph_mapping.xlsx"
    mapping_wb.save(mapping_path)

    with open(members_path, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("mph_members.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})
    with open(mapping_path, "rb") as f:
        client.post("/portfolio-analysis/subgroup-mapping/upload", files={"file": ("mph_mapping.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/master-clients")
    assert resp.json() == ["MPH CONSULTING SERVICES DMCC"]

    resp = client.get("/portfolio-analysis/summary", params={"group_by": "master_client"})
    assert resp.status_code == 200
    rows = resp.json()["rows"]
    assert len(rows) == 1
    assert rows[0]["master_client"] == "MPH CONSULTING SERVICES DMCC"
    assert rows[0]["member_count"] == len(subgroups)


def test_filter_options_returns_only_values_actually_present(client, members_two_networks_xlsx, mapping_xlsx, rate_card_xlsx):
    with open(members_two_networks_xlsx, "rb") as f:
        client.post("/portfolio-analysis/members/upload", files={"file": ("members.xlsx", f, "application/octet-stream")})
    with open(mapping_xlsx, "rb") as f:
        client.post("/portfolio-analysis/group-product-mapping/upload", files={"file": ("mapping.xlsx", f, "application/octet-stream")})
    with open(rate_card_xlsx, "rb") as f:
        client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})

    resp = client.get("/portfolio-analysis/filter-options")
    assert resp.status_code == 200
    options = resp.json()
    assert options["network"] == ["MSH Comprehensive", "MSH Platinum"]
    assert options["gender"] == ["F", "M"]
    assert options["product"] == ["Gold"]
