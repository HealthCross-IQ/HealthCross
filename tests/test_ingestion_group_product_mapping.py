import openpyxl
import pytest

from app.ingestion.group_product_mapping import parse_group_product_mapping, parse_subgroup_master_mapping


@pytest.fixture()
def mapping_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Master Group", "Product"])
    ws.append(["Acme Sub LLC", "Acme Holdings", "Bronze"])
    # A row naming only the master group (no distinct sub-group) still
    # produces a usable mapping entry.
    ws.append([None, "Umbrella Co", "Gold"])
    path = tmp_path / "mapping.xlsx"
    wb.save(path)
    return path


def test_produces_one_row_per_populated_group_column(mapping_xlsx):
    with open(mapping_xlsx, "rb") as f:
        rows = parse_group_product_mapping(f, "mapping.xlsx")
    by_name = {r["group_name"]: r["product"] for r in rows}
    assert by_name == {
        "Acme Sub LLC": "Bronze",
        "Acme Holdings": "Bronze",
        "Umbrella Co": "Gold",
    }


def test_skips_rows_with_no_product(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Master Group", "Product"])
    ws.append(["Acme Sub LLC", "Acme Holdings", None])
    path = tmp_path / "mapping2.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_group_product_mapping(f, "mapping2.xlsx")
    assert rows == []


def test_parse_subgroup_master_mapping_pairs_each_subgroup_with_its_master(mapping_xlsx):
    with open(mapping_xlsx, "rb") as f:
        rows = parse_subgroup_master_mapping(f, "mapping.xlsx")
    # Only the row with BOTH a subgroup and a master group produces a pairing -
    # the master-only row ("Umbrella Co") has no subgroup to pair it with.
    assert rows == [{"subgroup_name": "Acme Sub LLC", "master_name": "Acme Holdings", "source_filename": "mapping.xlsx"}]


def test_parse_subgroup_master_mapping_skips_rows_missing_either_side(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Master Group", "Product"])
    ws.append(["Acme Sub LLC", None, "Bronze"])
    ws.append([None, "Umbrella Co", "Gold"])
    path = tmp_path / "mapping3.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_subgroup_master_mapping(f, "mapping3.xlsx")
    assert rows == []
