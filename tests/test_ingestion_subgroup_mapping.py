import openpyxl
import pytest

from app.ingestion.subgroup_mapping import parse_subgroup_mapping


@pytest.fixture()
def kiko_mapping_xlsx(tmp_path):
    # The real-world example the user shared - one master ("KIKO MIDDLE
    # EAST FZ LLC") with several subgroups, including the master itself
    # appearing as its own subgroup row.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Group Name"])
    ws.append(["BEAUTY COSMETICS W.L.L.", "KIKO MIDDLE EAST FZ LLC"])
    ws.append(["KIKO COSMETICS BAHRAIN WLL", "KIKO MIDDLE EAST FZ LLC"])
    ws.append(["KIKO COSMETICS JORDAN LLC", "KIKO MIDDLE EAST FZ LLC"])
    ws.append(["KIKO COSMETICS KUWAIT WLL", "KIKO MIDDLE EAST FZ LLC"])
    ws.append(["KIKO COSMETICS LLC", "KIKO MIDDLE EAST FZ LLC"])
    ws.append(["KIKO IMEA FZE", "KIKO MIDDLE EAST FZ LLC"])
    ws.append(["KIKO MIDDLE EAST FZ LLC", "KIKO MIDDLE EAST FZ LLC"])
    path = tmp_path / "kiko_mapping.xlsx"
    wb.save(path)
    return path


def test_parses_every_subgroup_to_its_master_group_name(kiko_mapping_xlsx):
    with open(kiko_mapping_xlsx, "rb") as f:
        rows = parse_subgroup_mapping(f, "kiko_mapping.xlsx")
    by_subgroup = {r["subgroup_name"]: r["master_name"] for r in rows}
    assert by_subgroup == {
        "BEAUTY COSMETICS W.L.L.": "KIKO MIDDLE EAST FZ LLC",
        "KIKO COSMETICS BAHRAIN WLL": "KIKO MIDDLE EAST FZ LLC",
        "KIKO COSMETICS JORDAN LLC": "KIKO MIDDLE EAST FZ LLC",
        "KIKO COSMETICS KUWAIT WLL": "KIKO MIDDLE EAST FZ LLC",
        "KIKO COSMETICS LLC": "KIKO MIDDLE EAST FZ LLC",
        "KIKO IMEA FZE": "KIKO MIDDLE EAST FZ LLC",
        "KIKO MIDDLE EAST FZ LLC": "KIKO MIDDLE EAST FZ LLC",
    }
    assert len(rows) == 7


def test_skips_rows_missing_either_column(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Group Name"])
    ws.append(["Acme Sub LLC", None])
    ws.append([None, "Acme Holdings"])
    path = tmp_path / "mapping2.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_subgroup_mapping(f, "mapping2.xlsx")
    assert rows == []


def test_accepts_master_group_column_alias(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Subgroup", "Master Group"])
    ws.append(["Acme Sub LLC", "Acme Holdings"])
    path = tmp_path / "mapping3.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_subgroup_mapping(f, "mapping3.xlsx")
    assert rows == [{"subgroup_name": "Acme Sub LLC", "master_name": "Acme Holdings", "source_filename": "mapping3.xlsx"}]


def test_finds_header_past_a_blank_title_row_and_blank_leading_columns(tmp_path):
    # The real file the user uploaded: a blank first row, then the header
    # starting in column D (three blank leading columns) rather than A1.
    # pd.read_excel's default header=0 previously saw only blank/"Unnamed"
    # columns here and rejected the whole file as empty.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None, None, None, None, None])
    ws.append([None, None, None, "Subgroup", "Group Name"])
    ws.append([None, None, None, "ACQUISIT DMCC", "ACQUISIT DMCC"])
    ws.append([None, None, None, "ACUMA LLC", "DEVERE ACUMA INSURANCE BROKERS LLC"])
    ws.append([None, None, None, "ACUMA LLC BRANCH", "DEVERE ACUMA INSURANCE BROKERS LLC"])
    path = tmp_path / "real_layout_mapping.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_subgroup_mapping(f, "real_layout_mapping.xlsx")
    by_subgroup = {r["subgroup_name"]: r["master_name"] for r in rows}
    assert by_subgroup == {
        "ACQUISIT DMCC": "ACQUISIT DMCC",
        "ACUMA LLC": "DEVERE ACUMA INSURANCE BROKERS LLC",
        "ACUMA LLC BRANCH": "DEVERE ACUMA INSURANCE BROKERS LLC",
    }
