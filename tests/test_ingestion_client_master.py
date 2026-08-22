from datetime import date

import openpyxl
import pytest

from app.ingestion.client_master import parse_client_master


@pytest.fixture()
def client_master_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Client Name (Master)", "OPEX", "Product", "Start Date"])
    ws.append(["CARL ZEISS MEDITEC AG", 27.5, "Gold", "2025-05-01"])
    ws.append(["ELLISDON CONSTRUCTION INC", 0.27, "Platinum", "2026-05-01"])
    ws.append(["DEVERE ACUMA INSURANCE BROKERS LLC", "20%", "Silver", None])
    path = tmp_path / "client_master.xlsx"
    wb.save(path)
    return path


def test_parses_opex_regardless_of_how_it_was_typed(client_master_xlsx):
    with open(client_master_xlsx, "rb") as f:
        rows = parse_client_master(f, "client_master.xlsx")
    by_name = {r["master_client_name"]: r for r in rows}
    # 27.5 (bare percent number), 0.27 (already a fraction), and "20%"
    # (percent string) should all normalize to a fraction consistently.
    assert by_name["CARL ZEISS MEDITEC AG"]["opex_pct"] == 0.275
    assert by_name["ELLISDON CONSTRUCTION INC"]["opex_pct"] == 0.27
    assert by_name["DEVERE ACUMA INSURANCE BROKERS LLC"]["opex_pct"] == 0.2


def test_parses_product_and_start_date(client_master_xlsx):
    with open(client_master_xlsx, "rb") as f:
        rows = parse_client_master(f, "client_master.xlsx")
    by_name = {r["master_client_name"]: r for r in rows}
    assert by_name["CARL ZEISS MEDITEC AG"]["product"] == "Gold"
    assert by_name["CARL ZEISS MEDITEC AG"]["start_date"] == date(2025, 5, 1)
    assert by_name["DEVERE ACUMA INSURANCE BROKERS LLC"]["start_date"] is None


def test_skips_rows_missing_the_client_name(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Client Name (Master)", "OPEX"])
    ws.append([None, 0.25])
    path = tmp_path / "client_master2.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_client_master(f, "client_master2.xlsx")
    assert rows == []


def test_accepts_loading_as_a_column_alias(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Master Client Name", "Loading"])
    ws.append(["Acme Holdings", 0.28])
    path = tmp_path / "client_master3.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_client_master(f, "client_master3.xlsx")
    assert rows == [{
        "master_client_name": "Acme Holdings", "opex_pct": 0.28,
        "product": None, "start_date": None, "end_date": None, "source_filename": "client_master3.xlsx",
    }]


def test_parses_end_date_and_supports_multiple_dated_rows_for_the_same_client(tmp_path):
    # A client whose real loading changed between renewals - two rows
    # for the SAME client, each its own Start Date/End Date window.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Client Name (Master)", "OPEX", "Start Date", "End Date"])
    ws.append(["Acme Holdings", 0.20, "2025-01-01", "2025-12-31"])
    ws.append(["Acme Holdings", 0.28, "2026-01-01", "2026-12-31"])
    path = tmp_path / "client_master4.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_client_master(f, "client_master4.xlsx")
    assert len(rows) == 2
    assert rows[0]["opex_pct"] == 0.20
    assert rows[0]["start_date"] == date(2025, 1, 1)
    assert rows[0]["end_date"] == date(2025, 12, 31)
    assert rows[1]["opex_pct"] == 0.28
    assert rows[1]["end_date"] == date(2026, 12, 31)
