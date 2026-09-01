"""API tests for the New Business rate-card pricing endpoints
(app/api/routes_new_business_rating.py) - uploading the two rate-card
spreadsheets, discovering priced options, and computing/storing a quote
for a case. Uses small synthetic spreadsheets built in-test rather than
HealthCross's real (commercially sensitive) rate card.
"""
import openpyxl
import pytest

PRODUCT_PRICING_HEADER = [
    "Product Name", "From Age", "To Age", "Male Price", "Female Price",
    "Married Female Price", "Region", "Network", "TPA", "Zone", "Created Date", "Updated Date",
]

VARIANT_HEADER = [
    "Benefit Name", "Variant Name", "Option Value", "Direction", "Impact Type",
    "Impact Value", "Is Default", "Zone", "Region", "TPA", "Network", "Created Date", "Updated Date",
]


def _write_xlsx(tmp_path, name, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    path = tmp_path / name
    wb.save(path)
    return path


@pytest.fixture()
def rate_card_files(tmp_path):
    pricing_path = _write_xlsx(
        tmp_path,
        "pricing.xlsx",
        PRODUCT_PRICING_HEADER,
        [
            ["Bronze", 0, 17, 1000, 1000, "Not Applicable", "Dubai", "Net A", "TPA X", "Worldwide", "2025-01-01", ""],
            ["Bronze", 18, 40, 2000, 2200, "0 (Applicable only for age band 18-50)", "Dubai", "Net A", "TPA X", "Worldwide", "2025-01-01", ""],
            ["Platinum", 18, 40, 5000, 5500, "0 (Applicable only for age band 18-50)", "Dubai", "Net A", "TPA X", "Worldwide", "2025-01-01", ""],
        ],
    )
    variants_path = _write_xlsx(
        tmp_path,
        "variants.xlsx",
        VARIANT_HEADER,
        [
            ["UAE Benefit", "Annual Limit", "USD 150,000", "Base", "Text", 0, "No", "Worldwide", "Dubai", "TPA X", "Net A", "2025-01-01", ""],
            ["UAE Benefit", "Annual Limit", "USD 500,000", "Upgrade", "Percent", 3, "No", "Worldwide", "Dubai", "TPA X", "Net A", "2025-01-01", ""],
        ],
    )
    return pricing_path, variants_path


def _upload_rate_cards(client, rate_card_files):
    pricing_path, variants_path = rate_card_files
    with open(pricing_path, "rb") as f:
        resp = client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200
    with open(variants_path, "rb") as f:
        resp = client.post("/admin/benefit-variant-rates/upload", files={"file": ("variants.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200


def _make_case(client, target_premium=None):
    db = client.db_session_local()
    from app.models import db_models as models

    case = models.Case(broker_name="Broker", company_name="Acme", industry="trading", target_premium=target_premium)
    db.add(case)
    db.commit()
    db.refresh(case)
    db.add_all(
        [
            models.CensusRecord(case_id=case.id, category="A", age=30, gender="M", marital_status="single", relation="employee", emirates="Dubai"),
            models.CensusRecord(case_id=case.id, category="A", age=28, gender="F", marital_status="married", relation="spouse", emirates="Dubai"),
        ]
    )
    db.commit()
    case_id = case.id
    db.close()
    return case_id


def test_upload_rate_card_replaces_existing_rows(client, rate_card_files):
    pricing_path, _ = rate_card_files
    with open(pricing_path, "rb") as f:
        resp = client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200
    body = resp.json()
    assert body["rows_ingested"] == 3
    assert set(body["products"]) == {"Bronze", "Platinum"}

    # A second upload should wholesale-replace, not append.
    with open(pricing_path, "rb") as f:
        resp = client.post("/admin/rate-cards/upload", files={"file": ("pricing.xlsx", f, "application/octet-stream")})
    assert resp.json()["rows_ingested"] == 3


def test_rate_card_options_lists_products_and_their_networks(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    resp = client.get("/new-business/rate-card-options")
    assert resp.status_code == 200
    body = resp.json()
    assert set(body["products"]) == {"Bronze", "Platinum"}
    assert body["product_networks"]["Bronze"] == [{"network": "Net A", "tpa": "TPA X"}]


def test_variant_options_returns_options_grouped_by_variant(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    resp = client.get("/new-business/variant-options", params={"region": "Dubai", "tpa": "TPA X", "network": "Net A"})
    assert resp.status_code == 200
    body = resp.json()
    values = {opt["option_value"] for opt in body["Annual Limit"]}
    assert values == {"USD 150,000", "USD 500,000"}


def test_compute_new_business_quote_for_a_case(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client, target_premium=5000)

    resp = client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    assert resp.status_code == 200
    body = resp.json()
    # net = 2000 (male) + 2200 (married female, Dubai has no surcharge) = 4200
    # loading = 10% + 5% + 5% + 6.5% (Bronze) = 26.5%
    assert body["result"]["categories"][0]["net_annual_premium"] == 4200.0
    assert body["case_gross_annual_premium"] == round(4200.0 / (1 - 0.28), 2)
    assert body["opportunity_assessment"]["verdict"] in {"Good", "Marginal", "Poor"}


def test_quote_persists_and_is_retrievable_as_latest(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)

    client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    resp = client.get(f"/cases/{case_id}/new-business-quote")
    assert resp.status_code == 200
    assert resp.json()["case_gross_annual_premium"] > 0

    resp = client.get(f"/cases/{case_id}/new-business-quotes")
    assert len(resp.json()) == 1


def test_quote_without_target_premium_reports_unknown_opportunity(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client, target_premium=None)

    resp = client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    assert resp.json()["opportunity_assessment"]["verdict"] == "Unknown"


def test_quote_missing_case_404s(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    resp = client.post(
        "/cases/999999/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    assert resp.status_code == 404


def test_quote_without_uploaded_rate_card_400s(client):
    case_id = _make_case(client)
    resp = client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    assert resp.status_code == 400


def test_census_categories_reports_member_counts_and_uncategorized(client):
    case_id = _make_case(client)
    db = client.db_session_local()
    from app.models import db_models as models

    db.add(models.CensusRecord(case_id=case_id, category=None, age=10, gender="M", marital_status="single", relation="child", emirates="Dubai"))
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/census-categories")
    assert resp.status_code == 200
    body = resp.json()
    assert body["categories"] == [{"category": "A", "member_count": 2}]
    assert body["uncategorized_member_count"] == 1


def test_census_categories_merges_inconsistently_cased_or_padded_category_values(client):
    # Regression test: a real case had "A" (44 members) and "A " (1 member,
    # a stray trailing space from the source spreadsheet) show up as two
    # separate Category A cards on the New Business Quote tab instead of
    # one merged 45-member category. Older rows uploaded before
    # app/ingestion/census.py started normalizing on parse can still carry
    # this raw, inconsistent data - the read side must merge them too.
    case_id = _make_case(client)
    db = client.db_session_local()
    from app.models import db_models as models

    db.query(models.CensusRecord).filter_by(case_id=case_id).delete()
    db.add_all(
        [
            models.CensusRecord(case_id=case_id, category="A", age=30, gender="M", marital_status="single", relation="employee", emirates="Dubai"),
            models.CensusRecord(case_id=case_id, category="A ", age=28, gender="F", marital_status="married", relation="spouse", emirates="Dubai"),
            models.CensusRecord(case_id=case_id, category="a", age=5, gender="M", marital_status="single", relation="child", emirates="Dubai"),
        ]
    )
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/census-categories")
    assert resp.status_code == 200
    assert resp.json()["categories"] == [{"category": "A", "member_count": 3}]


def test_census_categories_suggests_a_product_tier_from_the_existing_insurer(client):
    from app.models import db_models as models

    db = client.db_session_local()
    db.add(models.InsurerTierPreference(insurer_name="Allianz", suggested_product="Platinum"))
    case = models.Case(broker_name="Broker", company_name="Acme", industry="trading", existing_insurer="Allianz")
    db.add(case)
    db.commit()
    db.refresh(case)
    db.add(models.CensusRecord(case_id=case.id, category="A", age=30, gender="M", marital_status="single", relation="employee", emirates="Dubai"))
    db.commit()
    case_id = case.id
    db.close()

    resp = client.get(f"/cases/{case_id}/census-categories")
    assert resp.status_code == 200
    assert resp.json()["suggested_product"] == "Platinum"


def test_census_categories_suggested_product_is_none_for_an_unmapped_insurer(client):
    case_id = _make_case(client)
    db = client.db_session_local()
    from app.models import db_models as models

    case = db.get(models.Case, case_id)
    case.existing_insurer = "Some Insurer Not In The Table"
    db.commit()
    db.close()

    resp = client.get(f"/cases/{case_id}/census-categories")
    assert resp.status_code == 200
    assert resp.json()["suggested_product"] is None


def test_quote_includes_a_tier_ladder_per_category(client, tmp_path):
    pricing_path = _write_xlsx(
        tmp_path,
        "pricing_ladder.xlsx",
        PRODUCT_PRICING_HEADER,
        [
            ["Bronze", 18, 40, 2000, 2200, "0 (Applicable only for age band 18-50)", "Dubai", "MSH Regular", "MSH MENA", "Worldwide", "2025-01-01", ""],
            ["Silver", 18, 40, 3000, 3300, "0 (Applicable only for age band 18-50)", "Dubai", "MSH Premium", "MSH MENA", "Worldwide", "2025-01-01", ""],
            ["Silver", 18, 40, 3200, 3500, "0 (Applicable only for age band 18-50)", "Dubai", "MSH Enhanced", "MSH MENA", "Worldwide", "2025-01-01", ""],
        ],
    )
    with open(pricing_path, "rb") as f:
        resp = client.post("/admin/rate-cards/upload", files={"file": ("pricing_ladder.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200

    case_id = _make_case(client)
    resp = client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "MSH Regular", "tpa": "MSH MENA", "variant_selections": {}}]},
    )
    assert resp.status_code == 200
    body = resp.json()
    ladder = body["result"]["categories"][0]["tier_ladder"]
    products_in_ladder = [row["product"] for row in ladder]
    assert products_in_ladder == ["Silver", "Bronze"]  # Bronze is the floor - nothing below it

    silver = next(r for r in ladder if r["product"] == "Silver")
    # Every MSH network offered under Silver shows up, not just one.
    assert {n["network"] for n in silver["networks"]} == {"MSH Premium", "MSH Enhanced"}
    premium_row = next(n for n in silver["networks"] if n["network"] == "MSH Premium")
    assert premium_row["net_annual_premium"] == 3000.0 + 3300.0  # both census members priced under Silver too

    bronze = next(r for r in ladder if r["product"] == "Bronze")
    chosen_row = next(n for n in bronze["networks"] if n["is_chosen"])
    assert chosen_row["network"] == "MSH Regular"


def _insert_existing_plan_with_nb_pick(client, case_id, category, product=None, network=None, tpa=None):
    db = client.db_session_local()
    from app.models import db_models as models

    plan = models.BenefitPlan(
        case_id=case_id, role="existing", plan_name=f"Category {category}",
        category=category, nb_product=product, nb_network=network, nb_tpa=tpa,
        standard_summary={},
    )
    db.add(plan)
    db.commit()
    db.refresh(plan)
    plan_id = plan.id
    db.close()
    return plan_id


def test_benefits_category_pick_lets_the_first_quote_auto_compute_on_census_upload(client, rate_card_files, tmp_path):
    # Before any manual "Compute quote" ever happens, a case whose Benefits
    # tab category already has its own Product/Network/TPA set (see
    # BenefitPlan.nb_product/nb_network/nb_tpa) should get its first quote
    # automatically the moment a matching census is uploaded.
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)
    # _make_case's own census is category A - remove it so this test
    # controls exactly when the census (the auto-quote trigger) arrives.
    db = client.db_session_local()
    from app.models import db_models as models

    db.query(models.CensusRecord).filter_by(case_id=case_id).delete()
    db.commit()
    db.close()

    _insert_existing_plan_with_nb_pick(client, case_id, "A", product="Bronze", network="Net A", tpa="TPA X")

    resp = client.get(f"/cases/{case_id}/new-business-quote")
    assert resp.status_code == 404  # nothing yet - no census

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Employee Ref", "Category", "Age", "Gender", "Marital Status", "Relation", "Emirate"])
    ws.append(["E1", "A", 30, "M", "single", "employee", "Dubai"])
    ws.append(["E2", "A", 28, "F", "married", "spouse", "Dubai"])
    path = tmp_path / "census.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        resp = client.post(f"/cases/{case_id}/census", files={"file": ("census.xlsx", f, "application/octet-stream")})
    assert resp.status_code == 200

    resp = client.get(f"/cases/{case_id}/new-business-quote")
    assert resp.status_code == 200
    assert resp.json()["case_gross_annual_premium"] > 0
    quoted_category = resp.json()["categories"][0]
    assert quoted_category["product"] == "Bronze"
    assert quoted_category["network"] == "Net A"
    assert quoted_category["tpa"] == "TPA X"


def test_without_a_benefits_category_pick_or_a_prior_quote_nothing_auto_computes(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)  # census category A has no matching benefit plan/pick

    resp = client.get(f"/cases/{case_id}/new-business-quote")
    assert resp.status_code == 404


def test_auto_requote_prefers_the_benefits_tab_pick_over_a_stale_prior_quote(client, rate_card_files):
    # The Benefits tab is the source of truth going forward - if it's
    # updated after an earlier manual quote, a later auto re-quote should
    # pick up the new value rather than keep reusing the old one.
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)

    resp = client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Platinum", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    assert resp.status_code == 200

    _insert_existing_plan_with_nb_pick(client, case_id, "A", product="Bronze", network="Net A", tpa="TPA X")

    # Trigger via a real re-upload of the census.
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Employee Ref", "Category", "Age", "Gender", "Marital Status", "Relation", "Emirate"])
    ws.append(["E1", "A", 30, "M", "single", "employee", "Dubai"])
    ws.append(["E2", "A", 28, "F", "married", "spouse", "Dubai"])

    import io

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(f"/cases/{case_id}/census", files={"file": ("census.xlsx", buf, "application/octet-stream")})
    assert resp.status_code == 200

    resp = client.get(f"/cases/{case_id}/new-business-quotes")
    quotes = resp.json()
    assert len(quotes) == 2  # the manual quote, plus one auto-triggered by the census re-upload
    latest = quotes[0]  # ordered newest first
    assert latest["categories"][0]["product"] == "Bronze"  # the Benefits tab's own pick won, not the stale prior quote


def test_saving_a_categorys_nb_pick_on_the_benefits_tab_auto_computes_the_quote_immediately(client, rate_card_files):
    # Saving Product/Network/TPA on a Benefits tab category card is itself
    # the input that unlocks auto-quoting - the underwriter shouldn't have
    # to also re-upload the census just to see the quote appear.
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)  # already has a census, per _make_case
    plan_id = _insert_existing_plan_with_nb_pick(client, case_id, "A")  # no product/network/tpa yet

    resp = client.get(f"/cases/{case_id}/new-business-quote")
    assert resp.status_code == 404

    resp = client.put(
        f"/cases/{case_id}/benefits/{plan_id}/summary",
        json={"fields": {}, "nb_product": "Bronze", "nb_network": "Net A", "nb_tpa": "TPA X"},
    )
    assert resp.status_code == 200

    resp = client.get(f"/cases/{case_id}/new-business-quote")
    assert resp.status_code == 200
    assert resp.json()["categories"][0]["product"] == "Bronze"


def test_new_business_quote_by_tier_covers_every_product_and_case_totals(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)

    resp = client.post(
        f"/cases/{case_id}/new-business-quote",
        json={"categories": [{"category": "A", "product": "Bronze", "network": "Net A", "tpa": "TPA X", "variant_selections": {}}]},
    )
    assert resp.status_code == 200

    resp = client.get(f"/cases/{case_id}/new-business-quote/by-tier")
    assert resp.status_code == 200
    by_tier = resp.json()
    assert [r["product"] for r in by_tier] == ["Platinum", "Gold", "Silver", "Bronze"]
    bronze = next(r for r in by_tier if r["product"] == "Bronze")
    platinum = next(r for r in by_tier if r["product"] == "Platinum")
    assert bronze["case_gross_annual_premium"] == pytest.approx(round(4200.0 / (1 - 0.28), 2), rel=1e-6)
    assert platinum["case_gross_annual_premium"] > bronze["case_gross_annual_premium"]


def test_new_business_quote_by_tier_404s_without_a_prior_quote(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)

    resp = client.get(f"/cases/{case_id}/new-business-quote/by-tier")
    assert resp.status_code == 404


def _insert_stale_quote(client, case_id, category_letter):
    # Simulates a quote persisted BEFORE category normalization existed -
    # its own stored `categories`/`result` carry the category as typed at
    # the time (e.g. lowercase "a"), never rewritten since. Regression
    # coverage for a real bug: /by-tier and /burning-cost-comparison
    # re-price live against the CURRENT (now-normalized) census, so a stale
    # un-normalized category here must still be matched, not silently
    # produce zero members/premium for every category despite the quote's
    # own stored total being real.
    db = client.db_session_local()
    from app.models import db_models as models

    result = {
        "categories": [
            {
                "category": category_letter, "product": "Bronze", "network": "Net A", "tpa": "TPA X",
                "member_count": 2, "net_annual_premium": 4200.0, "loading_pct": 0.28,
                "gross_annual_premium": 5714.29, "member_breakdown": [], "warnings": [],
            }
        ],
        "case_gross_annual_premium": 5714.29,
        "priced_member_count": 2,
        "uncategorized_member_count": 0,
    }
    quote = models.NewBusinessQuote(
        case_id=case_id,
        categories=[
            {"category": category_letter, "product": "Bronze", "network": "Net A", "tpa": "TPA X", "commission_pct": None, "variant_selections": {}}
        ],
        case_gross_annual_premium=5714.29,
        result=result,
    )
    db.add(quote)
    db.commit()
    db.close()


def test_auto_requote_resolves_from_a_stale_un_normalized_prior_quote(client, rate_card_files):
    # No Benefits-tab pick exists for this category - the auto re-quote
    # can only resolve Product/Network/TPA from the prior quote, whose own
    # category was stored un-normalized ("a") from before the fix. Must
    # still match the (now-normalized) census category "A".
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)  # census category "A", 2 members
    _insert_stale_quote(client, case_id, "a")

    import io

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Employee Ref", "Category", "Age", "Gender", "Marital Status", "Relation", "Emirate"])
    ws.append(["E1", "A", 30, "M", "single", "employee", "Dubai"])
    ws.append(["E2", "A", 28, "F", "married", "spouse", "Dubai"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = client.post(f"/cases/{case_id}/census", files={"file": ("census.xlsx", buf, "application/octet-stream")})
    assert resp.status_code == 200

    resp = client.get(f"/cases/{case_id}/new-business-quotes")
    quotes = resp.json()
    assert len(quotes) == 2  # the stale quote, plus one auto-triggered by the census re-upload
    latest = quotes[0]
    assert latest["categories"][0]["category"] == "A"
    assert latest["categories"][0]["product"] == "Bronze"
    assert latest["case_gross_annual_premium"] > 0


def test_new_business_quote_by_tier_matches_a_stale_un_normalized_category_from_before_the_fix(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)  # census category "A" (2 members)
    _insert_stale_quote(client, case_id, "a")  # stored lowercase, never re-saved since

    resp = client.get(f"/cases/{case_id}/new-business-quote/by-tier")
    assert resp.status_code == 200
    bronze = next(r for r in resp.json() if r["product"] == "Bronze")
    assert bronze["case_gross_annual_premium"] == pytest.approx(round(4200.0 / (1 - 0.28), 2), rel=1e-6)
    assert bronze["categories"][0]["member_count"] == 2


# --- GET /cases/{id}/opportunity-assessment -----------------------------

def _upload_portfolio_book(client):
    """A minimal book with claims, so there is experience to assess
    against - the assessment reads every benchmark off it.
    """
    from datetime import date

    from app.models import db_models as models

    db = client.db_session_local()
    members, claims = [], []
    for i in range(60):
        beneficiary = f"B{i}"
        relation = "Employee" if i % 3 else "Child"
        members.append(dict(
            beneficiary_id=beneficiary, relation=relation,
            age=35 if relation == "Employee" else (0 if i % 6 == 0 else 10),
            gender="M", nationality_zone="Zone 1",
            policy_start_date=date(2025, 1, 1), policy_end_date=date(2026, 1, 1),
            member_start_date=date(2025, 1, 1), member_end_date=date(2026, 1, 1),
            gross_premium=10_000.0, net_premium=8_000.0,
        ))
        claims.append(dict(
            patient_id=beneficiary, final_amount=5_000.0, claim_status="Paid Claims",
            date_of_treatment=date(2025, 6, 1), ip_op_maternity="OP",
            policy_start_date=date(2025, 1, 1), policy_end_date=date(2026, 1, 1),
            member_start_date=date(2025, 1, 1), member_end_date=date(2026, 1, 1),
        ))
    db.bulk_insert_mappings(models.PortfolioMember, members)
    db.bulk_insert_mappings(models.PortfolioClaimEntry, claims)
    db.commit()
    db.close()


def test_opportunity_assessment_needs_a_census_first(client):
    db = client.db_session_local()
    from app.models import db_models as models

    case = models.Case(broker_name="B", company_name="C", industry="trading")
    db.add(case)
    db.commit()
    case_id = case.id
    db.close()

    resp = client.get(f"/cases/{case_id}/opportunity-assessment")
    assert resp.status_code == 400
    assert "census" in resp.json()["detail"].lower()


def test_opportunity_assessment_says_so_when_there_is_no_book_to_assess_against(client):
    case_id = _make_case(client)
    resp = client.get(f"/cases/{case_id}/opportunity-assessment")
    assert resp.status_code == 400
    assert "experience" in resp.json()["detail"].lower()


def test_opportunity_assessment_returns_factors_and_one_conclusion(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    _upload_portfolio_book(client)
    case_id = _make_case(client)

    resp = client.get(f"/cases/{case_id}/opportunity-assessment")
    assert resp.status_code == 200
    body = resp.json()
    assert body["factors"], "an assessment with no factors is not an assessment"
    assert body["verdict"]["verdict"]
    assert body["required_margin_pct"] >= 0
    # Every factor states how it should be treated, and nothing that the
    # cube already prices is allowed to move the number.
    assert all(f["treatment"] in
               {"already_in_price", "load", "widen_margin", "ask"} for f in body["factors"])
    moving = {c["key"] for c in body["required_margin_contributions"]}
    priced_in = {f["key"] for f in body["factors"] if f["treatment"] == "already_in_price"}
    assert not (moving & priced_in)


def test_opportunity_assessment_carries_the_open_questions_it_cannot_answer(client, rate_card_files):
    _upload_rate_cards(client, rate_card_files)
    _upload_portfolio_book(client)
    case_id = _make_case(client)

    body = client.get(f"/cases/{case_id}/opportunity-assessment").json()
    assert {q["key"] for q in body["open_questions"]} == {
        "participation", "incumbent_loss_ratio", "reason_for_moving"
    }


# --- the issued quote outranks the rate card's variants ------------------

def _insert_quoted_plan_document(client, case_id, category, summary):
    from app.models import db_models as models

    db = client.db_session_local()
    db.add(models.BenefitPlan(
        case_id=case_id, role="quoted", category=category,
        plan_name=f"Gold - CAT {category}", source_format="pdf",
        standard_summary=summary,
    ))
    db.commit()
    db.close()


def test_the_issued_quote_document_supplies_the_proposed_side(client, rate_card_files):
    # A benefit the card prices as part of the product rather than as a
    # dropdown reads as "not priced as a variant" - but it IS quoted, and
    # stated, in the document the broker received. Values off the real
    # Haworth quote PDF.
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)
    _insert_quoted_plan_document(client, case_id, "A", {
        "health_screening_wellness": "USD 500 Once per policy year",
        "maternity_limit": "USD 14,000",
        "area_of_cover": "Worldwide Excluding USA",
    })

    rows = {
        r["field"]: r
        for r in client.get(f"/cases/{case_id}/existing-vs-proposed").json()["categories"][0]["rows"]
    }
    assert rows["health_screening_wellness"]["proposed"] == "USD 500 Once per policy year"
    assert rows["maternity_limit"]["proposed"] == "USD 14,000"


def test_a_field_the_quote_parser_could_not_read_does_not_blank_the_card_value(client, rate_card_files):
    # The parser returns "Not specified in source document" for a field
    # it could not find. Letting that through would replace an answer the
    # rate card DID resolve with an apology.
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)
    _insert_quoted_plan_document(client, case_id, "A", {
        "annual_limit": "Not specified in source document",
        "dental": "USD 2,000",
    })

    rows = {
        r["field"]: r
        for r in client.get(f"/cases/{case_id}/existing-vs-proposed").json()["categories"][0]["rows"]
    }
    assert rows["annual_limit"]["proposed"] != "Not specified in source document"
    assert rows["dental"]["proposed"] == "USD 2,000"


# --- GET /cases/{id}/price-comparison -----------------------------------

def test_price_comparison_reads_the_issued_premium_off_the_quote_document(client, rate_card_files):
    from app.models import db_models as models

    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)
    db = client.db_session_local()
    db.add(models.BenefitPlan(
        case_id=case_id, role="quoted", category="A", plan_name="Gold - CAT A",
        gross_premium=179_192.0, member_count=9, standard_summary={},
    ))
    db.commit()
    db.close()

    body = client.get(f"/cases/{case_id}/price-comparison").json()
    assert body["prices"]["issued_price"] == 179_192.0
    assert body["issued_quote"]["categories_priced"] == 1


def test_price_comparison_without_an_issued_quote_reports_no_discount(client, rate_card_files):
    # A case nobody has issued yet has not been discounted by 100%.
    _upload_rate_cards(client, rate_card_files)
    case_id = _make_case(client)
    body = client.get(f"/cases/{case_id}/price-comparison").json()
    assert body["prices"]["issued_price"] is None
    assert body["discount"] is None
