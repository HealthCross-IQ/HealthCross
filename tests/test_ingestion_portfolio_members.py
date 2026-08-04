"""Tests for app/ingestion/portfolio_members.py - parses HealthCross's own
book-wide membership export. Uses a small synthetic spreadsheet (same
fixed column layout as the real export) rather than the real book, which
is real client/employee data and not committed to the repo.
"""
import openpyxl
import pytest

from app.ingestion.portfolio_members import parse_portfolio_members

HEADER = [
    "CONTRACT", "MASTERCONTRACT", "POLICYNUMBER", "MSH_POLICYNUMBER", "BENEFICIARYID",
    "DOB", "GENDER", "MARITALSTATUS", "NATIONALITY", "DEPENDENCY",
    "PERSONRESIDENCEEMIRATE", "CATEGORY", "NETWORKTYPE",
    "Eff Date", "Exp Date", "EndoDate (Member Start Date)", "EndoDate (Member End Date)",
    "GrossPremium", "ActualGrossPremium", "NETPREMIUM", "ACTUALNETPREMIUM", "TPA FEE",
]


@pytest.fixture()
def members_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    ws.append([
        "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001",
        "1990-01-01", "M", "Married", "India", "Principal",
        "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        12000, 12000, None, None, 500,
    ])
    ws.append([
        "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001C1",
        "2020-06-15", "F", "Single", "India", "Child",
        "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        2000, 2000, None, None, 100,
    ])
    path = tmp_path / "members.xlsx"
    wb.save(path)
    return path


def test_parses_member_fields(members_xlsx):
    with open(members_xlsx, "rb") as f:
        rows = parse_portfolio_members(f, "members.xlsx")
    assert len(rows) == 2

    employee = next(r for r in rows if r["beneficiary_id"] == "ACM0001")
    assert employee["contract"] == "Acme Sub LLC"
    assert employee["master_contract"] == "Acme Holdings"
    assert employee["network_type_raw"] == "PLATINUM"
    assert employee["relation"] == "employee"
    assert employee["gender"] == "M"
    assert employee["nationality"] == "India"
    assert employee["nationality_zone"] == "zone_1_asia"
    assert employee["region"] == "Dubai"
    assert employee["actual_gross_premium"] == 12000
    # This fixture's HEADER has neither column - an older-format export
    # still parses fine, just without these two (see the dedicated test
    # below for when they're present).
    assert employee["master_client_name"] is None
    assert employee["product_name"] is None


def test_reads_master_client_name_and_product_name_when_present(tmp_path):
    # Starting Aug 2026 underwriting adds these two columns directly into
    # the export itself (Column K / Column BD on a real file) instead of
    # maintaining them as separate Group->Product/Subgroup->Master mapping
    # uploads - see resolve_group_product/resolve_master_client in
    # app/scoring/rules/portfolio_analysis.py, which prefer these over the
    # uploaded mappings whenever present.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER + ["Master Client Name", "PRODUCTNAME"])
    ws.append([
        "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", "ACM0001",
        "1990-01-01", "M", "Married", "India", "Principal",
        "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        12000, 12000, None, None, 500,
        "Acme Holdings Group", "Gold",
    ])
    path = tmp_path / "members_with_product.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        rows = parse_portfolio_members(f, "members_with_product.xlsx")
    assert rows[0]["master_client_name"] == "Acme Holdings Group"
    assert rows[0]["product_name"] == "Gold"


def test_derives_age_from_dob_not_a_separate_age_column(members_xlsx):
    # The real export's own AGE column is always blank - age must come
    # from DOB instead.
    with open(members_xlsx, "rb") as f:
        rows = parse_portfolio_members(f, "members.xlsx")
    child = next(r for r in rows if r["beneficiary_id"] == "ACM0001C1")
    assert child["age"] is not None
    assert child["relation"] == "child"


def test_skips_rows_with_no_beneficiary_id(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    ws.append([
        "Acme Sub LLC", "Acme Holdings", "P100", "QC1-ACM-A", None,
        "1990-01-01", "M", "Married", "India", "Principal",
        "Dubai", "QIC/HC/BR/ACM/DXB/A", "PLATINUM",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        12000, 12000, None, None, 500,
    ])
    path = tmp_path / "members2.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_portfolio_members(f, "members2.xlsx")
    assert rows == []
