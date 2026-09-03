"""Tests for app/ingestion/portfolio_claims.py - parses HealthCross's own
book-wide claims export. Uses a small synthetic .xlsx spreadsheet (same
column shape as the real .xlsb export) since pyxlsb can only read .xlsb
files, not write them - the xlsb-specific engine dispatch is verified
separately via monkeypatching rather than a real binary fixture.
"""
import openpyxl
import pytest

from app.ingestion.portfolio_claims import parse_portfolio_claims

HEADER = [
    "PATIENT_ID", "CLAIM_ID", "Claim Status", "GROUP_NAME", "CLIENT_NAME", "MSH_POLICY_NUMBER",
    "POLICY_START_DATE", "POLICY_END_DATE", "Member Start Date", "Member End Date",
    "DATE_OF_TREATMENT", "DATE_RECEPTION", "RELATION", "IP_OP_MATERNITY", "MEDICAL_CATEGORY",
    "PROVIDER_NAME", "DIAGNOSIS_CODE", "DIAGNOSIS_DESCRIPTION", "Claimed Amount AED",
    "Final Amount in AED",
]


@pytest.fixture()
def claims_xlsx(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    ws.append([
        "ACM0001", "CLM1", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        "2025-06-01", "2025-06-08", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
        "J309", "Allergic rhinitis", 100.0, 90.0,
    ])
    ws.append([
        "ACM0001", "CLM2", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        "2025-07-01", "2025-07-03", "Main Insured", "IP", "SURGERY", "Some Hospital",
        "K358", "Appendicitis", 5000.0, 4800.0,
    ])
    path = tmp_path / "claims.xlsx"
    wb.save(path)
    return path


def test_parses_claim_fields(claims_xlsx):
    with open(claims_xlsx, "rb") as f:
        rows = parse_portfolio_claims(f, "claims.xlsx")
    assert len(rows) == 2
    first = rows[0]
    assert first["patient_id"] == "ACM0001"
    assert first["group_name"] == "Acme Sub LLC"
    assert first["client_name"] == "Acme Holdings"
    assert first["msh_policy_number"] == "QC1-ACM-A"
    assert first["final_amount"] == 90.0
    assert first["date_of_treatment"].isoformat() == "2025-06-01"
    # The date the claim actually reached the TPA - distinct from
    # date_of_treatment, and the pair a real completion-factor IBNR is
    # built on. Previously dropped on the floor entirely.
    assert first["date_reception"].isoformat() == "2025-06-08"


def test_a_claim_with_no_reception_date_yet_is_none_not_missing(tmp_path):
    # A freshly-treated claim that has not been submitted at all yet has
    # no reception date - that absence IS the IBNR signal, so it must
    # come back as None rather than being silently dropped or defaulted.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(HEADER)
    ws.append([
        "ACM0002", "CLM3", "Outstanding Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        "2026-08-30", None, "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
        "J309", "Allergic rhinitis", 50.0, 45.0,
    ])
    path = tmp_path / "claims_no_reception.xlsx"
    wb.save(path)
    with open(path, "rb") as f:
        rows = parse_portfolio_claims(f, "claims_no_reception.xlsx")
    assert rows[0]["date_reception"] is None


def test_reads_xlsb_files_with_the_calamine_engine(monkeypatch, claims_xlsx):
    import pandas as pd

    captured = {}
    real_read_excel = pd.read_excel

    def _spy_read_excel(file, **kwargs):
        captured["engine"] = kwargs.get("engine")
        # Reuse the real xlsx fixture's content - only the engine kwarg
        # actually matters for this test, not real xlsb bytes.
        with open(claims_xlsx, "rb") as f:
            return real_read_excel(f, **kwargs)

    monkeypatch.setattr(pd, "read_excel", _spy_read_excel)
    with open(claims_xlsx, "rb") as f:
        parse_portfolio_claims(f, "claims.xlsb")
    assert captured["engine"] == "calamine"


def test_reads_xlsx_with_the_calamine_engine(monkeypatch, claims_xlsx):
    import pandas as pd

    captured = {}
    real_read_excel = pd.read_excel

    def _spy_read_excel(file, **kwargs):
        captured["engine"] = kwargs.get("engine")
        return real_read_excel(file, **kwargs)

    monkeypatch.setattr(pd, "read_excel", _spy_read_excel)
    with open(claims_xlsx, "rb") as f:
        parse_portfolio_claims(f, "claims.xlsx")
    assert captured["engine"] == "calamine"


@pytest.fixture()
def claims_xlsx_with_leading_blank_sheet(tmp_path):
    # A real HealthCross export shipped exactly this shape once: a
    # completely blank "Sheet1" ahead of the actual data, which now lives
    # on a separately named "Sheet 1" - reading only the first sheet
    # silently parsed 0 rows instead of erroring (see parse_portfolio_claims).
    wb = openpyxl.Workbook()
    blank = wb.active
    blank.title = "Sheet1"
    data = wb.create_sheet("Sheet 1")
    data.append(HEADER)
    data.append([
        "ACM0001", "CLM1", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        "2025-06-01", "2025-06-08", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
        "J309", "Allergic rhinitis", 100.0, 90.0,
    ])
    path = tmp_path / "claims_with_blank_sheet.xlsx"
    wb.save(path)
    return path


def test_skips_a_leading_blank_sheet_to_find_the_real_data(claims_xlsx_with_leading_blank_sheet):
    with open(claims_xlsx_with_leading_blank_sheet, "rb") as f:
        rows = parse_portfolio_claims(f, "claims.xlsx")
    assert len(rows) == 1
    assert rows[0]["patient_id"] == "ACM0001"


@pytest.fixture()
def claims_xlsx_with_pivot_summary_sheets(tmp_path):
    # The real HealthCross_Claims export ships pivot/summary sheets AHEAD
    # of the per-claim-line data: "Premium", "Claims", "LOSS RATIO",
    # "Loading" (each a pivot whose real headers sit below a title row, so
    # pandas names their columns "Unnamed: N"), a partial "Detail1"
    # extract, and the full "DATA" sheet. Taking the first non-empty sheet
    # picked up a pivot summary and silently produced all-None rows.
    wb = openpyxl.Workbook()
    premium = wb.active
    premium.title = "Premium"
    premium.append([None, None, None, None])
    premium.append([None, "ONE Group", "Eff Date", "Total"])
    premium.append([None, "ACQUISIT DMCC", "2025-05-01", 883707.16])

    detail = wb.create_sheet("Detail1")  # a partial extract, same headers as DATA
    detail.append(HEADER)
    detail.append([
        "PAR0001", "CLMP", "Paid Claims", "Partial Sub", "Partial Holdings", "QC9-PAR-A",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        "2025-06-01", "2025-06-08", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
        "J309", "Allergic rhinitis", 10.0, 10.0,
    ])

    data = wb.create_sheet("DATA")  # the real, full claim-line sheet
    data.append(HEADER)
    for i in range(3):
        data.append([
            f"ACM000{i}", f"CLM{i}", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
            "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
            "2025-06-01", "2025-06-08", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
            "J309", "Allergic rhinitis", 100.0, 90.0,
        ])

    path = tmp_path / "claims_with_pivots.xlsx"
    wb.save(path)
    return path


def test_picks_the_claim_line_sheet_over_leading_pivot_summary_sheets(claims_xlsx_with_pivot_summary_sheets):
    with open(claims_xlsx_with_pivot_summary_sheets, "rb") as f:
        rows = parse_portfolio_claims(f, "claims.xlsb")

    # Must pick "DATA" (3 real claim lines), not the leading "Premium"
    # pivot (which would parse as all-None rows) nor the smaller
    # "Detail1" partial extract that shares DATA's own headers.
    assert len(rows) == 3
    assert {r["patient_id"] for r in rows} == {"ACM0000", "ACM0001", "ACM0002"}
    assert all(r["final_amount"] == 90.0 for r in rows)


def test_a_pivot_only_workbook_does_not_masquerade_as_claim_data(tmp_path):
    # Nothing qualifies as claim-line data - the parser falls back to the
    # first non-empty sheet rather than picking a pivot as if it were real,
    # leaving the resulting all-None rows for callers to reject.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Premium"
    ws.append([None, "ONE Group", "Eff Date", "Total"])
    ws.append([None, "ACQUISIT DMCC", "2025-05-01", 883707.16])
    path = tmp_path / "pivots_only.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        rows = parse_portfolio_claims(f, "claims.xlsb")
    assert all(r["final_amount"] is None for r in rows)


def test_recognizes_the_contract_wording_for_amount_columns(tmp_path):
    # A newer HealthCross export renamed "Claimed Amount AED"/"Final Amount
    # in AED" to "Claimed Amount Contract"/"Final Amount Contract" (still
    # AED throughout - its own CONTRACT_CURRENCY column confirms that) -
    # without this alias claimed_amount/final_amount silently came back
    # None for every single row instead of erroring.
    header = [h if h not in ("Claimed Amount AED", "Final Amount in AED") else
               {"Claimed Amount AED": "Claimed Amount Contract", "Final Amount in AED": "Final Amount Contract"}[h]
               for h in HEADER]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    ws.append([
        "ACM0001", "CLM1", "Paid Claims", "Acme Sub LLC", "Acme Holdings", "QC1-ACM-A",
        "2025-01-01", "2026-01-01", "2025-01-01", "2026-01-01",
        "2025-06-01", "2025-06-08", "Main Insured", "OP", "PHARMACY", "Some Pharmacy",
        "J309", "Allergic rhinitis", 100.0, 90.0,
    ])
    path = tmp_path / "claims_contract_wording.xlsx"
    wb.save(path)

    with open(path, "rb") as f:
        rows = parse_portfolio_claims(f, "claims.xlsx")
    assert rows[0]["claimed_amount"] == 100.0
    assert rows[0]["final_amount"] == 90.0
